"""
gui_launcher_3d_apple.py  —  Siraal Manufacturing Engine  v6.0
═══════════════════════════════════════════════════════════════════════════
  ▸ A P P L E   D E S I G N   E D I T I O N  —  Sidebar Navigation
  ▸ Left-rail vertical tab system (VS Code / macOS Settings style)
  ▸ Real glassmorphism   — PIL blur-behind compositing
  ▸ Radial glow FX       — orb lights, accents, status indicators
  ▸ SF Pro typography    — macOS-native · Windows fallback
  ▸ Larger legible fonts throughout
  ▸ 100 % original functionality preserved
═══════════════════════════════════════════════════════════════════════════
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import threading, os, queue, math, time, datetime, traceback
from pathlib import Path
import pythoncom

# ── PIL (glassmorphism) ───────────────────────────────────────────────────
try:
    from PIL import Image, ImageDraw, ImageFilter, ImageTk
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ── Optional runtime deps ─────────────────────────────────────────────────
try:    import pandas as pd;          PANDAS_OK  = True
except: PANDAS_OK  = False
try:    import openpyxl;              OPENPYXL_OK = True
except: OPENPYXL_OK = False
try:
    from watchdog.observers import Observer
    from watchdog.events    import FileSystemEventHandler
    WATCHDOG_OK = True
except: WATCHDOG_OK = False


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ①  DESIGN TOKENS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

_ROOT_BG = "#06050F"
_SURF    = "#0C0A1D"
_CARD    = "#111028"
_SIDEBAR = "#080716"

def _hex_rgb(h):
    h = h.lstrip("#")
    return int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)

def _blend(fg, a, bg=_ROOT_BG):
    f, b = _hex_rgb(fg), _hex_rgb(bg)
    r = [int(f[i]*a + b[i]*(1-a)) for i in range(3)]
    return "#{:02X}{:02X}{:02X}".format(*r)

BLUE   = "#0A84FF"; TEAL   = "#5AC8FA"; GREEN  = "#30D158"
ORANGE = "#FF9F0A"; RED    = "#FF453A"; PURPLE = "#BF5AF2"
YELLOW = "#FFD60A"; PINK   = "#FF375F"

TEXT_1 = "#F2F1FC"
TEXT_2 = _blend("#F2F1FC", 0.62)
TEXT_3 = _blend("#F2F1FC", 0.36)

GLASS_BASE = _blend("#FFFFFF", 0.055, _SURF)
BORDER_S   = _blend("#FFFFFF", 0.07,  _SURF)
BORDER_M   = _blend("#FFFFFF", 0.15,  _SURF)

C = {
    "void":_ROOT_BG,"base":_SURF,"surface":GLASS_BASE,"elevated":_blend("#FFFFFF",0.09,_SURF),
    "card":_CARD,"border":BORDER_S,"border2":BORDER_M,
    "gold":YELLOW,"gold_dim":_blend(YELLOW,0.38),"amber":ORANGE,
    "teal":TEAL,"teal_dim":_blend(TEAL,0.38),"violet":PURPLE,
    "violet_dim":_blend(PURPLE,0.38),"ok":GREEN,"warn":ORANGE,
    "error":RED,"info":BLUE,"text":TEXT_1,"text2":TEXT_2,"text3":TEXT_3,
}

# ── Typography ────────────────────────────────────────────────────────────
import tkinter.font as _tkf
def _pick(*fams):
    try:
        av = set(_tkf.families())
        for f in fams:
            if f in av: return f
    except Exception: pass
    return fams[-1]

_FF = _pick("SF Pro Display","Helvetica Neue",".SF NS Display","Segoe UI","Helvetica")
_FT = _pick("SF Pro Text",   "Helvetica Neue",".SF NS Text",   "Segoe UI","Helvetica")
_FM = _pick("SF Mono","JetBrains Mono","Cascadia Code","Menlo","Consolas")

def FM(s): return (_FM, s)

GEAR_TYPES = ["Spur_Gear_3D","Helical_Gear","Ring_Gear_3D","Bevel_Gear","Worm","Worm_Wheel"]
MATERIALS  = ["Steel-4140","Steel-1020","Al-6061","Brass-C360","Nylon-66","Ti-6Al-4V"]
PRIORITIES = ["High","Medium","Low"]

def _profile_shift(Z, PA=20.0):
    a = math.radians(PA)
    z_min = 2.0 / math.sin(a)**2
    return round(max(0.0,(z_min-Z)/z_min),4) if Z < z_min else 0.0


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ②  PIL — Background & Glassmorphism
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _make_bg(w, h):
    if not PIL_OK or w < 2 or h < 2: return None
    base  = Image.new("RGBA", (w, h), (*_hex_rgb(_ROOT_BG), 255))
    layer = Image.new("RGBA", (w, h), (0,0,0,0))
    orbs  = [
        (w*0.15, h*0.12, 460, (10,132,255), 34),
        (w*0.88, h*0.08, 360, (191,90,242), 28),
        (w*0.55, h*0.85, 500, (10,132,255), 22),
        (w*0.88, h*0.65, 320, (90,200,250), 20),
        (w*0.04, h*0.80, 240, (191,90,242), 16),
        (w*0.50, h*0.04, 280, (48,209,88),  14),
    ]
    for cx, cy, rad, rgb, max_a in orbs:
        r2  = int(rad)
        orb = Image.new("RGBA", (r2*2+4, r2*2+4), (0,0,0,0))
        d   = ImageDraw.Draw(orb)
        for ri in range(r2, 0, -2):
            alpha = int(max_a*(ri/r2)**1.9)
            d.ellipse([r2-ri, r2-ri, r2+ri, r2+ri], fill=(*rgb, alpha))
        orb = orb.filter(ImageFilter.GaussianBlur(radius=max(1, r2//6)))
        px, py = int(cx-r2-2), int(cy-r2-2)
        cx0, cy0 = max(0,-px), max(0,-py)
        crop = orb.crop((cx0, cy0, orb.width, orb.height))
        layer.paste(crop, (max(0,px), max(0,py)), crop)
    return Image.alpha_composite(base, layer).convert("RGB")


def _make_glass(bg, x, y, w, h, tint="#FFFFFF", tint_a=0.07, blur=18):
    if not PIL_OK or bg is None or w < 4 or h < 4: return None
    iw, ih = bg.size
    x2, y2 = min(x+w, iw), min(y+h, ih)
    x,  y  = max(0, x),    max(0, y)
    if x >= x2 or y >= y2: return None
    crop = bg.crop((x, y, x2, y2))
    if crop.size != (w, h):
        pad = Image.new("RGB", (w, h), _hex_rgb(_ROOT_BG)); pad.paste(crop); crop = pad
    blurred  = crop.filter(ImageFilter.GaussianBlur(radius=blur)).convert("RGBA")
    tr,tg,tb = _hex_rgb(tint)
    tint_img = Image.new("RGBA", (w,h), (tr,tg,tb, int(255*tint_a)))
    glass    = Image.alpha_composite(blurred, tint_img)
    d = ImageDraw.Draw(glass)
    d.rectangle([0,0,w,0],     fill=(255,255,255,95))
    d.rectangle([0,1,w,1],     fill=(255,255,255,40))
    d.rectangle([0,0,0,h],     fill=(255,255,255,15))
    d.rectangle([w-1,0,w-1,h], fill=(0,0,0,40))
    d.rectangle([0,h-1,w,h-1], fill=(0,0,0,40))
    return glass.convert("RGB")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ③  GLASS PANEL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class GlassPanel(ctk.CTkFrame):
    def __init__(self, parent, app, tint="#FFFFFF", tint_a=0.07,
                 accent=None, blur=18, **kwargs):
        kwargs.setdefault("fg_color",      GLASS_BASE)
        kwargs.setdefault("corner_radius", 14)
        kwargs.setdefault("border_width",  1)
        kwargs.setdefault("border_color",  BORDER_M)
        super().__init__(parent, **kwargs)
        self._app=app; self._tint=tint; self._tint_a=tint_a
        self._accent=accent; self._blur=blur
        self._photo=None; self._bg_lbl=None
        if hasattr(app, "_glass_panels"): app._glass_panels.append(self)
        self.bind("<Destroy>", self._unregister)

    def _unregister(self, _=None):
        if hasattr(self._app,"_glass_panels") and self in self._app._glass_panels:
            self._app._glass_panels.remove(self)

    def apply_glass(self):
        if not PIL_OK or not getattr(self._app,"_bg_pil",None): return
        self.update_idletasks()
        rx = self.winfo_rootx() - self._app.winfo_rootx()
        ry = self.winfo_rooty() - self._app.winfo_rooty()
        w, h = self.winfo_width(), self.winfo_height()
        if w < 4 or h < 4: return
        g = _make_glass(self._app._bg_pil, rx, ry, w, h,
                        self._tint, self._tint_a, self._blur)
        if g is None: return
        if self._accent:
            ar,ag,ab = _hex_rgb(self._accent)
            gra = g.convert("RGBA"); d = ImageDraw.Draw(gra)
            d.rectangle([0,0,w,1], fill=(ar,ag,ab,210))
            d.rectangle([0,2,w,2], fill=(ar,ag,ab,80))
            g = gra.convert("RGB")
        self._photo = ImageTk.PhotoImage(g)
        if self._bg_lbl is None:
            self._bg_lbl = tk.Label(self, image=self._photo, bd=0, highlightthickness=0)
            self._bg_lbl.place(x=0,y=0,relwidth=1,relheight=1)
            self._bg_lbl.lower()
        else:
            self._bg_lbl.configure(image=self._photo)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ④  SHARED WIDGET HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _divider(parent, label="", color=BORDER_M):
    fr = ctk.CTkFrame(parent, fg_color="transparent", height=26)
    fr.pack(fill="x", padx=12, pady=(10,2)); fr.pack_propagate(False)
    ctk.CTkFrame(fr, fg_color=color, height=1, corner_radius=0).place(
        relx=0, rely=0.5, relwidth=1.0, anchor="w")
    if label:
        lbl = ctk.CTkLabel(fr, text=f"  {label}  ",
                           font=ctk.CTkFont(_FT, 10, "bold"),
                           text_color=TEXT_3, fg_color=GLASS_BASE)
        lbl.place(relx=0.04, rely=0.5, anchor="w")


def _log_widget(parent):
    outer = ctk.CTkFrame(parent, fg_color=_blend(_ROOT_BG,1.0),
                         corner_radius=12, border_width=1, border_color=BORDER_S)
    txt = tk.Text(outer, bg=_blend("#040310",1.0), fg="#7DFF9A",
                  font=FM(10), relief="flat", padx=12, pady=10,
                  state="disabled", cursor="arrow", wrap="word",
                  selectbackground=BORDER_M)
    sb = ctk.CTkScrollbar(outer, command=txt.yview)
    txt.configure(yscrollcommand=sb.set)
    sb.pack(side="right", fill="y"); txt.pack(fill="both", expand=True)
    for tag,col in [("ok",GREEN),("warn",ORANGE),("err",RED),
                    ("info",BLUE),("head",YELLOW),("ai",PURPLE),("cost",_blend(PURPLE,0.85))]:
        txt.tag_config(tag, foreground=col)
    return outer, txt


def _append_log(txt, msg, tag=""):
    txt.configure(state="normal")
    ts   = datetime.datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}]  {msg}\n"
    if not tag:
        if   any(x in msg for x in ("✔","OK","Done","complete","COMPLETE")): tag="ok"
        elif any(x in msg for x in ("⚠","WARN","warn","WARNING")):           tag="warn"
        elif any(x in msg for x in ("✘","ERROR","error","fail","FAIL")):     tag="err"
        elif any(x in msg for x in ("SYSTEM","BOM","LOADING","SAVING")):     tag="info"
        elif any(x in msg for x in ("AI","Gemini","Designing","model")):     tag="ai"
        elif any(x in msg for x in ("PDF","Economic","ESG","Cost")):         tag="cost"
        elif msg.startswith(("╔","║","╚")):                                   tag="head"
    txt.insert("end", line, tag or "")
    txt.see("end"); txt.configure(state="disabled")


def _pill_btn(parent, text, cmd, color, width=None, height=34):
    kw = dict(height=height,
              fg_color=_blend(color,0.13), hover_color=_blend(color,0.24),
              border_color=color, border_width=1,
              text_color=color, font=ctk.CTkFont(_FT,11,"bold"),
              corner_radius=9, command=cmd)
    if width: kw["width"] = width
    return ctk.CTkButton(parent, text=text, **kw)


def _run_btn(parent, text, cmd, color, height=42):
    return ctk.CTkButton(parent, text=text, height=height,
                         fg_color=color, hover_color=_blend(color,0.80),
                         text_color="#FFFFFF", font=ctk.CTkFont(_FF,12,"bold"),
                         corner_radius=10, command=cmd)


def _entry(parent, var, placeholder="", height=32):
    return ctk.CTkEntry(parent, textvariable=var, height=height,
                        fg_color=_blend("#FFFFFF",0.06,_CARD), border_color=BORDER_M,
                        text_color=TEXT_1, font=ctk.CTkFont(_FT,11),
                        placeholder_text=placeholder)


def _combo(parent, var, values, height=32):
    return ctk.CTkComboBox(parent, variable=var, values=values, height=height,
                           fg_color=_blend("#FFFFFF",0.06,_CARD),
                           button_color=BORDER_M, border_color=BORDER_M,
                           text_color=TEXT_1, font=ctk.CTkFont(_FT,11))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑤  EXCEL WATCHER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class ExcelWatcher:
    def __init__(self, on_change):
        self._on_change=on_change; self._observer=None
        self._path=None; self._last_snap={}

    def _snapshot(self):
        if not (self._path and PANDAS_OK): return {}
        try:
            xl=pd.read_excel(self._path, sheet_name=None, dtype=str); snap={}
            for sname,df in xl.items(): snap[sname]=df.fillna("").to_dict(orient="records")
            return snap
        except Exception: return {}

    def start(self, path):
        self.stop(); self._path=path; self._last_snap=self._snapshot()
        if not WATCHDOG_OK: return
        class _H(FileSystemEventHandler):
            def __init__(self_,cb): self_.cb=cb
            def on_modified(self_,ev):
                if Path(ev.src_path).resolve()==Path(path).resolve(): self_.cb()
        self._observer=Observer()
        self._observer.schedule(_H(self._check), str(Path(path).parent), recursive=False)
        self._observer.start()

    def _check(self):
        time.sleep(0.4); new=self._snapshot(); diffs=[]
        for sheet in set(list(self._last_snap)+list(new)):
            o=self._last_snap.get(sheet,[]); n=new.get(sheet,[])
            if o!=n:
                diffs.append(f"Sheet [{sheet}]: {abs(len(n)-len(o))} row delta")
                for i,(oo,nn) in enumerate(zip(o,n)):
                    for k in set(list(oo)+list(nn)):
                        ov,nv=oo.get(k,""),nn.get(k,"")
                        if ov!=nv: diffs.append(f"  Row {i+1} [{k}]  {ov!r}→{nv!r}")
        if diffs: self._last_snap=new; self._on_change(diffs)

    def stop(self):
        if self._observer:
            try: self._observer.stop(); self._observer.join(timeout=1)
            except Exception: pass
            self._observer=None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑥  EXCEL WRITER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def write_bom_excel(parts, out_path):
    if not OPENPYXL_OK: raise ImportError("openpyxl not installed")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title="BOM_Gears"
    FH="0D3A5C"; FS="1A2B3C"; FA="1C2D3E"; FW="3A2000"
    GC="F0B429"; WC="E8EDF2"; GR="8B99AA"
    thin=Side(style="thin",color="2E3E52"); bord=Border(left=thin,right=thin,top=thin,bottom=thin)
    def hf(h): return PatternFill("solid",fgColor=h)

    # Row 1 — decorative title (validator skips this, header=2 in pandas)
    ws.merge_cells("A1:P1"); c=ws["A1"]
    c.value="SIRAAL 3D GEAR ENGINE — BILL OF MATERIALS"
    c.font=Font(name="Segoe UI",bold=True,color=GC,size=11)
    c.fill=hf(FH); c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=22

    # Row 2 — subtitle (validator skips this too)
    ws.merge_cells("A2:P2"); c2=ws["A2"]
    c2.value="IS 2535 / ISO 1328  |  Involute 20°  |  TN-IMPACT 2026  |  1st Angle Projection"
    c2.font=Font(name="Segoe UI",color=GR,size=8); c2.fill=hf(FS)
    c2.alignment=Alignment(horizontal="center"); ws.row_dimensions[2].height=14

    # Row 3 — REAL column headers (validator reads these as pandas header=2)
    COLS=["#","Part_Number","Part_Type","Material","Param_1","Param_2","Param_3","Param_4",
          "Qty","Priority","Enabled","Description","Mass_kg","Est_Cost_Rs","Notes","Profile_Shift_x"]
    WIDTHS=[4,20,16,12,10,10,10,10,5,8,7,30,8,12,30,10]
    for ci,(col,w) in enumerate(zip(COLS,WIDTHS),1):
        cell=ws.cell(row=3,column=ci,value=col)
        cell.font=Font(name="Segoe UI",bold=True,color=WC,size=9)
        cell.fill=hf("1A3A5C"); cell.border=bord
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=28

    # Rows 4+ — part data
    for ri,p in enumerate(parts,4):
        pt=p.get("Part_Type","Spur_Gear_3D"); Z=int(float(p.get("Param_1",20)))
        m=float(p.get("Param_2",3)); fw=float(p.get("Param_3",30)); bdd=float(p.get("Param_4",20))
        mat=p.get("Material","Steel-4140")
        x=_profile_shift(Z) if pt in ("Spur_Gear_3D","Helical_Gear","Worm_Wheel") else 0.0
        ra=Z*m/2+m*(1+x); rb=bdd/2
        vol=math.pi*(ra**2-max(rb,0)**2)*fw*0.92
        DENS={"Steel-1020":7.87e-3,"Steel-4140":7.85e-3,"Al-6061":2.70e-3,
              "Brass-C360":8.50e-3,"Nylon-66":1.14e-3,"Ti-6Al-4V":4.43e-3}
        COST={"Steel-1020":125,"Steel-4140":185,"Al-6061":265,
              "Brass-C360":520,"Nylon-66":415,"Ti-6Al-4V":3800}
        mass=round(vol*DENS.get(mat,7.85e-3)/1e6,3); cost=round(mass*COST.get(mat,185),2)
        rf=hf(FW if x>0 else (FA if ri%2==0 else FS))
        vals=[ri-3,p.get("Part_Number",""),pt,mat,Z,m,fw,bdd,
              p.get("Quantity",1),p.get("Priority","High"),"YES",
              p.get("Description",""),mass,cost,p.get("Notes",""),x]
        for ci,v in enumerate(vals,1):
            cell=ws.cell(row=ri,column=ci,value=v); cell.fill=rf; cell.border=bord
            cell.font=Font(name="Segoe UI",color=WC,size=9)
            cell.alignment=Alignment(vertical="center",
                horizontal="center" if ci in (1,5,6,7,8,9,10,11,16) else "left")
        ws.row_dimensions[ri].height=16

    # Totals footer
    tr=4+len(parts); ws.merge_cells(f"A{tr}:D{tr}")
    c=ws.cell(row=tr,column=1,value="TOTALS (all enabled):")
    c.font=Font(name="Segoe UI",bold=True,color=GC,size=9)
    c.fill=hf(FH); c.alignment=Alignment(horizontal="right")
    for ci in range(1,17): ws.cell(row=tr,column=ci).fill=hf(FH)

    ws.freeze_panes="B4"; wb.save(out_path); return out_path


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑦  GEAR ROW
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class GearRow(ctk.CTkFrame):
    _counter = 0
    def __init__(self, parent, on_delete, on_change, main_app):
        super().__init__(parent, fg_color=_CARD, corner_radius=10,
                         border_width=1, border_color=BORDER_S)
        GearRow._counter += 1
        self.main_app=main_app; self._on_delete=on_delete; self._on_change=on_change
        self.pack(fill="x", pady=(0,6)); self._build()

    def _field(self, parent, label, var, width=80, choices=None):
        fr = ctk.CTkFrame(parent, fg_color="transparent"); fr.pack(side="left", padx=(0,6))
        ctk.CTkLabel(fr, text=label, font=ctk.CTkFont(_FT,9), text_color=TEXT_3).pack(anchor="w")
        FG = _blend("#FFFFFF",0.06,_CARD)
        if label == "Type":
            w = ctk.CTkComboBox(fr, variable=var,
                                values=self.main_app._get_all_part_types(),
                                width=width, height=30,
                                fg_color=FG, border_color=BORDER_M, button_color=BORDER_M,
                                text_color=TEXT_1, font=ctk.CTkFont(_FT,10),
                                command=lambda _: self._on_change())
            w.bind("<Button-1>", lambda e: w.configure(values=self.main_app._get_all_part_types()))
            w.pack()
        elif choices:
            w = ctk.CTkComboBox(fr, variable=var, values=choices,
                                width=width, height=30,
                                fg_color=FG, border_color=BORDER_M, button_color=BORDER_M,
                                text_color=TEXT_1, font=ctk.CTkFont(_FT,10),
                                command=lambda _: self._on_change()); w.pack()
        else:
            w = ctk.CTkEntry(fr, textvariable=var, width=width, height=30,
                             fg_color=FG, border_color=BORDER_M,
                             text_color=TEXT_1, font=ctk.CTkFont(_FT,10)); w.pack()
            var.trace_add("write", lambda *_: self._on_change())
        return w

    def _build(self):
        n=GearRow._counter
        top=ctk.CTkFrame(self, fg_color="transparent"); top.pack(fill="x", padx=10, pady=(8,2))
        self.v_pno=ctk.StringVar(value=f"GR-{n:03d}"); self.v_type=ctk.StringVar(value="Spur_Gear_3D")
        self.v_mat=ctk.StringVar(value="Steel-4140"); self.v_z=ctk.StringVar(value="20")
        self.v_m=ctk.StringVar(value="3"); self.v_fw=ctk.StringVar(value="30")
        self.v_bd=ctk.StringVar(value="20"); self.v_qty=ctk.StringVar(value="1")
        self.v_prio=ctk.StringVar(value="High"); self.v_desc=ctk.StringVar(value="")
        self._field(top,"Part No",   self.v_pno,  100)
        self._field(top,"Type",      self.v_type, 140, None)
        self._field(top,"Material",  self.v_mat,  115, MATERIALS)
        self._field(top,"Z/Starts",  self.v_z,    60)
        self._field(top,"Module m",  self.v_m,    56)
        self._field(top,"Face W",    self.v_fw,   56)
        self._field(top,"Bore Ø",    self.v_bd,   56)
        self._field(top,"Qty",       self.v_qty,  44)
        self._field(top,"Priority",  self.v_prio, 88, PRIORITIES)
        tail=ctk.CTkFrame(top,fg_color="transparent"); tail.pack(side="left",padx=(6,0))
        ctk.CTkButton(tail,text="✕",width=30,height=30,
                      fg_color=_blend(RED,0.15,_CARD),hover_color=_blend(RED,0.28,_CARD),
                      text_color=RED,border_color=_blend(RED,0.5,_CARD),border_width=1,
                      corner_radius=7,font=ctk.CTkFont(_FF,10,"bold"),
                      command=self._on_delete).pack()
        dr=ctk.CTkFrame(self,fg_color="transparent"); dr.pack(fill="x",padx=10,pady=(0,8))
        ctk.CTkLabel(dr,text="Desc:",font=ctk.CTkFont(_FT,9),text_color=TEXT_3).pack(side="left",padx=(0,5))
        ctk.CTkEntry(dr,textvariable=self.v_desc,height=26,
                     fg_color=_blend("#FFFFFF",0.04,_CARD),border_color=BORDER_S,
                     text_color=TEXT_2,font=ctk.CTkFont(_FT,10)).pack(side="left",fill="x",expand=True)
        self._warn_lbl=ctk.CTkLabel(dr,text="",font=ctk.CTkFont(_FT,9),text_color=ORANGE)
        self._warn_lbl.pack(side="right",padx=(6,0))
        for v in (self.v_z,self.v_m,self.v_bd,self.v_fw):
            v.trace_add("write",lambda *_: self._live_validate())

    def _live_validate(self):
        try:
            Z=int(self.v_z.get()); m=float(self.v_m.get())
            bd=float(self.v_bd.get()); fw=float(self.v_fw.get())
            w=[]; x=_profile_shift(Z)
            if x>0:         w.append(f"⚠ x={x:.3f}")
            if Z<6:         w.append("⚠ Z<6")
            if bd/2>=Z*m/2: w.append("⚠ bore≥PCD/2")
            if fw/m>12:     w.append("⚠ fw/m>12")
            if fw/m<6:      w.append("⚠ fw/m<6")
            self._warn_lbl.configure(text="  ".join(w))
        except Exception: self._warn_lbl.configure(text="")

    def get_part(self):
        return {"Part_Number":self.v_pno.get(),"Part_Type":self.v_type.get(),
                "Material":self.v_mat.get(),"Param_1":self.v_z.get(),
                "Param_2":self.v_m.get(),"Param_3":self.v_fw.get(),
                "Param_4":self.v_bd.get(),"Qty":self.v_qty.get(),
                "Priority":self.v_prio.get(),"Description":self.v_desc.get(),"Enabled":"YES"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑧  SIDEBAR NAV BUTTON
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class SideNavBtn(ctk.CTkFrame):
    _H = 42   # fixed height for every nav item

    def __init__(self, parent, icon, label, color, on_click, **kwargs):
        kwargs["height"] = self._H
        super().__init__(parent, fg_color="transparent", cursor="hand2", **kwargs)
        self.pack_propagate(False)   # never let children stretch the height
        self._color=color; self._active=False; self._on_click=on_click

        # Accent bar on left edge
        self._bar = ctk.CTkFrame(self, fg_color="transparent", width=3, corner_radius=2)
        self._bar.pack(side="left", fill="y", padx=(3,0))
        self._bar.pack_propagate(False)

        # Icon + label — fill="x" only, no vertical expand
        inner = ctk.CTkFrame(self, fg_color="transparent")
        inner.pack(side="left", fill="x", expand=True, padx=(6,4))

        self._icon_lbl = ctk.CTkLabel(inner, text=icon,
                                       font=ctk.CTkFont(_FF,15), text_color=TEXT_3, width=28)
        self._icon_lbl.pack(side="left")
        self._text_lbl = ctk.CTkLabel(inner, text=label,
                                       font=ctk.CTkFont(_FF,12), text_color=TEXT_3, anchor="w")
        self._text_lbl.pack(side="left", fill="x", expand=True)

        for w in (self, inner, self._icon_lbl, self._text_lbl, self._bar):
            w.bind("<Button-1>", self._click)
        self.bind("<Enter>", self._hover_in)
        self.bind("<Leave>", self._hover_out)

    def _click(self, _=None): self._on_click()

    def _hover_in(self, _=None):
        if not self._active: self.configure(fg_color=_blend(self._color,0.10,_SIDEBAR))

    def _hover_out(self, _=None):
        if not self._active: self.configure(fg_color="transparent")

    def set_active(self, active):
        self._active = active
        if active:
            self.configure(fg_color=_blend(self._color,0.18,_SIDEBAR))
            self._bar.configure(fg_color=self._color)
            self._icon_lbl.configure(text_color=self._color)
            self._text_lbl.configure(text_color=TEXT_1, font=ctk.CTkFont(_FF,12,"bold"))
        else:
            self.configure(fg_color="transparent")
            self._bar.configure(fg_color="transparent")
            self._icon_lbl.configure(text_color=TEXT_3)
            self._text_lbl.configure(text_color=TEXT_3, font=ctk.CTkFont(_FF,12))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑨  MAIN WINDOW
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class SiraalGUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("SIRAAL Manufacturing Engine  |  TN-IMPACT 2026")
        self.geometry("1560x960"); self.minsize(1200,800)
        self.configure(fg_color=_ROOT_BG)

        self._q1=queue.Queue(); self._q2=queue.Queue()
        self._q3=queue.Queue(); self._q4=queue.Queue()
        self._watcher          = ExcelWatcher(self._on_excel_change)
        self._watch_active     = False
        self._gear_rows: list[GearRow] = []
        self._pending_copilot_data     = None
        self._glass_panels: list[GlassPanel] = []
        self._bg_pil   = None; self._bg_photo = None
        self._active_tab = None
        self._tab_frames = {}; self._nav_btns = {}

        self._build_bg_canvas()
        self._build_root_layout()
        self._build_sidebar()
        self._build_all_tabs()
        self._build_watcher_bar()

        self._switch_tab("excel")

        self.after(100, self._poll_q1); self.after(110, self._poll_q2)
        self.after(120, self._poll_q3); self.after(130, self._poll_q4)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.bind("<Configure>", self._on_resize)
        self.after(250, self._render_bg)

    # ── Background ─────────────────────────────────────────────────────────
    def _build_bg_canvas(self):
        self._bg_canvas = tk.Canvas(self, bg=_ROOT_BG, highlightthickness=0, bd=0)
        self._bg_canvas.place(x=0, y=0, relwidth=1, relheight=1)
        self._bg_canvas.tk.call("lower", self._bg_canvas._w)

    def _render_bg(self, _=None):
        w, h = self.winfo_width(), self.winfo_height()
        if w < 10 or h < 10: self.after(200, self._render_bg); return
        self._bg_pil = _make_bg(w, h)
        if self._bg_pil and PIL_OK:
            self._bg_photo = ImageTk.PhotoImage(self._bg_pil)
            self._bg_canvas.delete("all")
            self._bg_canvas.create_image(0, 0, image=self._bg_photo, anchor="nw")
        self._apply_all_glass()

    def _apply_all_glass(self):
        for p in list(self._glass_panels):
            try: p.apply_glass()
            except Exception: pass

    def _on_resize(self, event):
        if event.widget is self:
            try: self.after_cancel(self._resize_job)
            except Exception: pass
            self._resize_job = self.after(320, self._render_bg)

    # ── Root layout ────────────────────────────────────────────────────────
    def _build_root_layout(self):
        self._main_fr = ctk.CTkFrame(self, fg_color="transparent")
        self._main_fr.pack(fill="both", expand=True)
        self._main_fr.columnconfigure(1, weight=1); self._main_fr.rowconfigure(0, weight=1)

        self._sidebar_fr = ctk.CTkFrame(self._main_fr, fg_color=_SIDEBAR,
                                         corner_radius=0, border_width=0, width=215)
        self._sidebar_fr.grid(row=0, column=0, sticky="nsew")
        self._sidebar_fr.pack_propagate(False); self._sidebar_fr.grid_propagate(False)

        self._content_fr = ctk.CTkFrame(self._main_fr, fg_color=_SURF,
                                         corner_radius=0, border_width=0)
        self._content_fr.grid(row=0, column=1, sticky="nsew")
        self._content_fr.rowconfigure(0, weight=1); self._content_fr.columnconfigure(0, weight=1)

        ctk.CTkFrame(self._main_fr, fg_color=BORDER_S, width=1, corner_radius=0
                     ).grid(row=0, column=0, sticky="nse")

    # ── Sidebar ────────────────────────────────────────────────────────────
    def _build_sidebar(self):
        sb = self._sidebar_fr

        # Logo
        logo_fr = ctk.CTkFrame(sb, fg_color="transparent")
        logo_fr.pack(fill="x", padx=10, pady=(16,6))
        ctk.CTkLabel(logo_fr, text="⚙", font=ctk.CTkFont(_FF,26), text_color=BLUE
                     ).pack(side="left", padx=(6,8))
        tf = ctk.CTkFrame(logo_fr, fg_color="transparent"); tf.pack(side="left")
        ctk.CTkLabel(tf, text="SIRAAL", font=ctk.CTkFont(_FF,15,"bold"), text_color=TEXT_1
                     ).pack(anchor="w")
        ctk.CTkLabel(tf, text="Mfg Engine  v6.0", font=ctk.CTkFont(_FT,9), text_color=TEXT_3
                     ).pack(anchor="w")

        ctk.CTkFrame(sb, fg_color=BORDER_S, height=1, corner_radius=0
                     ).pack(fill="x", padx=8, pady=(8,10))

        # Nav items — tight spacing, fixed height enforced by SideNavBtn itself
        for key, icon, label, color in [
            ("excel",   "📂", "Excel Mode",       TEAL),
            ("manual",  "🔧", "Manual Mode",      YELLOW),
            ("ai",      "🤖", "AI Shape Creator", PURPLE),
            ("copilot", "🧠", "AI BOM Copilot",   BLUE),
        ]:
            btn = SideNavBtn(sb, icon, label, color,
                             on_click=lambda k=key: self._switch_tab(k))
            btn.pack(fill="x", padx=6, pady=2)
            self._nav_btns[key] = btn

        # Spacer
        ctk.CTkFrame(sb, fg_color="transparent").pack(fill="both", expand=True)

        ctk.CTkFrame(sb, fg_color=BORDER_S, height=1, corner_radius=0
                     ).pack(fill="x", padx=8, pady=(0,8))

        # ESG button
        self._btn_global_cost = ctk.CTkButton(
            sb, text="📊  ESG & COST",
            fg_color=_blend(PURPLE,0.20,_SIDEBAR), hover_color=_blend(PURPLE,0.32,_SIDEBAR),
            border_color=PURPLE, border_width=1, text_color=TEXT_1,
            font=ctk.CTkFont(_FF,11,"bold"), height=36, corner_radius=10,
            command=self._run_global_cost)
        self._btn_global_cost.pack(fill="x", padx=10, pady=(0,6))

        # Badges
        bf = ctk.CTkFrame(sb, fg_color="transparent"); bf.pack(fill="x", padx=8, pady=(0,12))
        for badge,col in [("IS 2535",TEAL),("ISO 1328",BLUE)]:
            ctk.CTkLabel(bf, text=badge, font=ctk.CTkFont(_FT,9,"bold"),
                         text_color=col, fg_color=_blend(col,0.14,_SIDEBAR),
                         corner_radius=5, padx=6, pady=2
                         ).pack(side="left", padx=3)

    # ── Tab management ─────────────────────────────────────────────────────
    def _switch_tab(self, key):
        for fr in self._tab_frames.values(): fr.grid_remove()
        for k, btn in self._nav_btns.items(): btn.set_active(k==key)
        if key in self._tab_frames:
            self._tab_frames[key].grid(row=0, column=0, sticky="nsew")
        self._active_tab = key
        self.after(80, self._apply_all_glass)

    def _build_all_tabs(self):
        for key in ("excel","manual","ai","copilot"):
            fr = ctk.CTkFrame(self._content_fr, fg_color="transparent")
            self._tab_frames[key] = fr
        self._build_excel_tab()
        self._build_manual_tab()
        self._build_ai_tab()
        self._build_copilot_tab()

    # ── Two-column scaffold ────────────────────────────────────────────────
    def _make_two_col(self, tab_key, left_w=295, accent=TEAL):
        fr = self._tab_frames[tab_key]
        fr.columnconfigure(0, weight=0, minsize=left_w)
        fr.columnconfigure(1, weight=1); fr.rowconfigure(0, weight=1)
        lp = GlassPanel(fr, self, tint="#FFFFFF", tint_a=0.06, accent=accent, width=left_w)
        lp.grid(row=0, column=0, sticky="nsew", padx=(10,6), pady=10)
        lp.pack_propagate(False)
        rp = GlassPanel(fr, self, tint="#FFFFFF", tint_a=0.05)
        rp.grid(row=0, column=1, sticky="nsew", padx=(0,10), pady=10)
        return lp, rp

    # ── Helpers ────────────────────────────────────────────────────────────
    def _update_tbl(self, widget, parts, states):
        widget.configure(state="normal"); widget.delete("1.0","end")
        widget.insert("end",f"{'PART NUMBER':<24} {'TYPE':<18} {'STATUS'}\n","head")
        widget.insert("end","─"*64+"\n","head")
        for p in parts:
            pno=p.get("Part_Number",""); pt=p.get("Part_Type","")
            st=states.get(pno,"⏳ Queued")
            line=f"{pno:<24} {pt:<18} {st}\n"
            if "✔" in st:   widget.insert("end",line,"ok")
            elif "✘" in st: widget.insert("end",line,"err")
            elif "⚙" in st: widget.insert("end",line,"warn")
            else:            widget.insert("end",line)
        widget.configure(state="disabled")

    def _clear_log(self, w):
        w.configure(state="normal"); w.delete("1.0","end"); w.configure(state="disabled")

    def _on_close(self): self._watcher.stop(); self.destroy()

    def _filter_manual_rows(self, *args):
        q=self._manual_search_var.get().lower()
        for row in self._gear_rows:
            if q in row.v_pno.get().lower() or q in row.v_type.get().lower():
                row.pack(fill="x",pady=(0,6))
            else: row.pack_forget()

    def _get_all_part_types(self):
        native=list(GEAR_TYPES); customs=[]
        tp=os.path.join(os.path.dirname(os.path.abspath(__file__)),"templates")
        if os.path.exists(tp):
            for f in os.listdir(tp):
                if f.startswith("Custom_") and f.endswith(".json"):
                    customs.append(f.replace(".json",""))
        return native+sorted(customs)

    def _get_custom_template_names(self):
        customs=[]; tp=os.path.join(os.path.dirname(os.path.abspath(__file__)),"templates")
        if os.path.exists(tp):
            for f in os.listdir(tp):
                if f.startswith("Custom_") and f.endswith(".json"):
                    customs.append(f.replace(".json",""))
        return sorted(customs)

    def _refresh_custom_buttons(self):
        for c in self._custom_btn_fr.winfo_children(): c.destroy()
        templates=self._get_custom_template_names()
        if not templates:
            ctk.CTkLabel(self._custom_btn_fr,
                         text="No AI templates yet — use AI Shape Creator.",
                         font=ctk.CTkFont(_FT,11,"italic"),text_color=TEXT_3
                         ).pack(pady=10,padx=10); return
        for name in templates:
            _pill_btn(self._custom_btn_fr, f"🤖 {name.replace('Custom_','')}",
                      lambda n=name: self._add_row(n), PURPLE, height=30).pack(side="left",padx=4)

    # ═══════════════════════════════════════════════════════════════════════
    # TAB 1 — EXCEL MODE
    # ═══════════════════════════════════════════════════════════════════════
    def _build_excel_tab(self):
        lp, rp = self._make_two_col("excel", left_w=300, accent=TEAL)
        # ── Left ──
        ctk.CTkFrame(lp,fg_color=TEAL,height=3,corner_radius=2).pack(fill="x")
        ctk.CTkLabel(lp,text="📂  EXCEL MODE",font=ctk.CTkFont(_FF,16,"bold"),text_color=TEAL
                     ).pack(pady=(18,3),padx=14)
        ctk.CTkLabel(lp,text="Load BOM xlsx · Validate · Generate 3D",
                     font=ctk.CTkFont(_FT,11),text_color=TEXT_3).pack(pady=(0,12))

        _divider(lp,"BOM FILE",TEAL)
        self._e1_file=ctk.StringVar(value="excels/demo_gears_3d.xlsx")
        _entry(lp,self._e1_file).pack(fill="x",padx=12,pady=(6,4))
        r=ctk.CTkFrame(lp,fg_color="transparent"); r.pack(fill="x",padx=12,pady=(0,4))
        _pill_btn(r,"Browse",self._browse_e1,TEAL,height=30).pack(side="left")

        _divider(lp,"FILTERS",TEAL)
        ctk.CTkLabel(lp,text="Part Type",font=ctk.CTkFont(_FT,11),text_color=TEXT_3
                     ).pack(anchor="w",padx=14,pady=(4,0))
        self._e1_type=ctk.StringVar(value="All")
        _combo(lp,self._e1_type,["All"]+GEAR_TYPES).pack(fill="x",padx=12,pady=(4,4))
        ctk.CTkLabel(lp,text="Priority",font=ctk.CTkFont(_FT,11),text_color=TEXT_3
                     ).pack(anchor="w",padx=14)
        self._e1_prio=ctk.StringVar(value="All")
        _combo(lp,self._e1_prio,["All"]+PRIORITIES).pack(fill="x",padx=12,pady=(4,12))

        _divider(lp,"ACTIONS",TEAL)
        _pill_btn(lp,"①  Validate BOM",self._validate_e1,TEAL,height=36).pack(fill="x",padx=12,pady=(8,5))
        self._btn_e1=_run_btn(lp,"②  GENERATE 3D PARTS",self._run_e1,TEAL)
        self._btn_e1.pack(fill="x",padx=12,pady=(0,8))

        sf=ctk.CTkFrame(lp,fg_color=_blend("#FFFFFF",0.04,_CARD),corner_radius=10)
        sf.pack(fill="x",padx=12,pady=(0,6))
        sr=ctk.CTkFrame(sf,fg_color="transparent"); sr.pack(fill="x",padx=10,pady=(8,4))
        self._e1_status_dot=ctk.CTkLabel(sr,text="●",font=ctk.CTkFont(_FF,13),text_color=GREEN)
        self._e1_status_dot.pack(side="left")
        self._e1_status=ctk.CTkLabel(sr,text="  Ready",font=ctk.CTkFont(_FF,13,"bold"),text_color=GREEN)
        self._e1_status.pack(side="left")
        self._e1_prog=ctk.CTkProgressBar(sf,fg_color=_blend("#FFFFFF",0.08,_CARD),
                                          progress_color=TEAL,height=6,corner_radius=3)
        self._e1_prog.set(0); self._e1_prog.pack(fill="x",padx=10,pady=(0,3))
        self._e1_prog_lbl=ctk.CTkLabel(sf,text="",font=ctk.CTkFont(_FT,10),text_color=TEXT_3)
        self._e1_prog_lbl.pack(pady=(0,8))
        info=ctk.CTkFrame(lp,fg_color=_blend("#FFFFFF",0.03,_CARD),corner_radius=8)
        info.pack(fill="x",padx=12,pady=(0,12))
        self._e1_stats=ctk.CTkLabel(info,text="No BOM loaded",font=FM(10),text_color=TEXT_3,justify="left")
        self._e1_stats.pack(padx=10,pady=8,anchor="w")

        # ── Right ──
        rp.rowconfigure(2,weight=1); rp.columnconfigure(0,weight=1)
        tbl_fr=ctk.CTkFrame(rp,fg_color=_blend("#FFFFFF",0.04,_SURF),
                            corner_radius=10,border_width=1,border_color=BORDER_S)
        tbl_fr.grid(row=0,column=0,sticky="ew",padx=10,pady=(10,6))
        ctk.CTkLabel(tbl_fr,text="  PART LIST",font=ctk.CTkFont(_FF,11,"bold"),text_color=TEAL
                     ).pack(anchor="w",padx=8,pady=(6,0))
        self._e1_tbl=tk.Text(tbl_fr,height=9,bg=_blend("#FFFFFF",0.04,_SURF),fg=TEXT_1,
                             font=FM(10),relief="flat",padx=10,pady=5,state="disabled",cursor="arrow")
        for tag,col in [("ok",GREEN),("err",RED),("warn",ORANGE),("head",TEAL)]:
            self._e1_tbl.tag_config(tag,foreground=col)
        sb2=ctk.CTkScrollbar(tbl_fr,command=self._e1_tbl.yview)
        self._e1_tbl.configure(yscrollcommand=sb2.set)
        sb2.pack(side="right",fill="y"); self._e1_tbl.pack(fill="both",expand=True)

        hdr_r=ctk.CTkFrame(rp,fg_color="transparent")
        hdr_r.grid(row=1,column=0,sticky="ew",padx=10,pady=(0,3))
        ctk.CTkLabel(hdr_r,text="  ⚙  ENGINE LOG",font=ctk.CTkFont(_FF,11,"bold"),text_color=TEAL).pack(side="left")
        _pill_btn(hdr_r,"Clear",lambda: self._clear_log(self._log1_txt),TEXT_3,height=26,width=60).pack(side="right",padx=6)
        log_fr,self._log1_txt=_log_widget(rp)
        log_fr.grid(row=2,column=0,sticky="nsew",padx=10,pady=(0,10))
        _append_log(self._log1_txt,"SYSTEM  Siraal Excel Mode — standby.","info")

    def _browse_e1(self):
        p=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xls"),("All","*.*")])
        if p: self._e1_file.set(p)

    def _validate_e1(self): threading.Thread(target=self._t_validate_e1,daemon=True).start()

    def _t_validate_e1(self):
        path=self._e1_file.get(); self._q1.put(("log",f"\n[BOM] VALIDATING {path}"))
        if not os.path.exists(path): self._q1.put(("log",f"✘ File not found: {path}")); return
        try:
            from validator_3d import Validator3D
            v=Validator3D(path,log_callback=lambda m: self._q1.put(("log",m)))
            v.run_checks(); parts=v.valid_parts
        except Exception as e: self._q1.put(("log",f"✘ {e}")); return
        pt=self._e1_type.get(); pr=self._e1_prio.get()
        if pt!="All": parts=[p for p in parts if p.get("Part_Type")==pt]
        if pr!="All": parts=[p for p in parts if p.get("Priority")==pr]
        self._q1.put(("stats",f"Parts: {len(parts)}\nFile: {os.path.basename(path)}"))
        self._q1.put(("table1",(parts,{p["Part_Number"]:"⏳ Queued" for p in parts})))
        col=GREEN if v.error_count==0 else ORANGE
        self._q1.put(("status1",(f"✔ {len(parts)} valid | {v.error_count} errors",col)))

    def _run_e1(self):
        self._btn_e1.configure(state="disabled")
        threading.Thread(target=self._t_run_e1,daemon=True).start()

    def _t_run_e1(self):
        path=self._e1_file.get()
        self._q1.put(("prog1",(0.05,"Validating…")))
        try:
            from validator_3d import Validator3D
            from autocad_engine_3d import AutoCAD3DGearEngine
            v=Validator3D(path,log_callback=lambda m: self._q1.put(("log",m)))
            v.run_checks(); parts=v.valid_parts
            pt=self._e1_type.get(); pr=self._e1_prio.get()
            if pt!="All": parts=[p for p in parts if p.get("Part_Type")==pt]
            if pr!="All": parts=[p for p in parts if p.get("Priority")==pr]
            if not parts:
                self._q1.put(("status1",("⚠ No parts after filters",ORANGE)))
                self._q1.put(("btn1",None)); return

            n = len(parts)
            st = {p["Part_Number"]: "⏳ Queued" for p in parts}
            self._q1.put(("table1",(parts,dict(st))))
            self._q1.put(("prog1",(0.15,"Starting AutoCAD…")))
            time.sleep(0.05)   # let poller tick before heavy work starts

            eng = AutoCAD3DGearEngine(log_cb=lambda m: self._q1.put(("log",m)))

            # ── One part at a time so every update is visible ───────────────
            for i, part in enumerate(parts):
                pno = part["Part_Number"]
                st[pno] = "⚙ Building…"
                pct = 0.20 + 0.75 * (i / n)
                self._q1.put(("table1",(parts,dict(st))))
                self._q1.put(("prog1",(pct, f"Building {pno}  ({i+1}/{n})")))
                time.sleep(0.05)   # yield so poller can repaint before build starts

                try:
                    eng.generate_3d_batch([part])
                    st[pno] = "✔ Done"
                except Exception as part_err:
                    st[pno] = "✘ Error"
                    self._q1.put(("log", f"✘ {pno}: {part_err}"))

                self._q1.put(("table1",(parts,dict(st))))
                self._q1.put(("prog1",(0.20 + 0.75*((i+1)/n),
                                       f"Done {pno}  ({i+1}/{n})")))

            done  = sum(1 for s in st.values() if s=="✔ Done")
            error = sum(1 for s in st.values() if s=="✘ Error")
            col   = GREEN if error==0 else ORANGE
            self._q1.put(("prog1",(1.0,"Complete!")))
            self._q1.put(("status1",(f"✔ {done} built  {f'| ✘ {error} errors' if error else ''}",col)))
        except Exception as e:
            self._q1.put(("log",f"✘ {e}\n{traceback.format_exc()}"))
            self._q1.put(("status1",("● Error",RED)))
        self._q1.put(("btn1",None))

    # ═══════════════════════════════════════════════════════════════════════
    # TAB 2 — MANUAL MODE
    # ═══════════════════════════════════════════════════════════════════════
    def _build_manual_tab(self):
        lp, rp = self._make_two_col("manual", left_w=290, accent=YELLOW)
        # ── Left ──
        ctk.CTkFrame(lp,fg_color=YELLOW,height=3,corner_radius=2).pack(fill="x")
        ctk.CTkLabel(lp,text="🔧  MANUAL MODE",font=ctk.CTkFont(_FF,16,"bold"),text_color=YELLOW
                     ).pack(pady=(18,3),padx=14)
        ctk.CTkLabel(lp,text="Type gear specs → Excel → 3D",
                     font=ctk.CTkFont(_FT,11),text_color=TEXT_3).pack(pady=(0,12))

        _divider(lp,"OUTPUT",YELLOW)
        ctk.CTkLabel(lp,text="Save Excel to:",font=ctk.CTkFont(_FT,11),text_color=TEXT_3
                     ).pack(anchor="w",padx=14,pady=(4,0))
        self._m_out=ctk.StringVar(value="output_manual/manual_bom.xlsx")
        _entry(lp,self._m_out).pack(fill="x",padx=12,pady=(4,4))
        _pill_btn(lp,"Choose…",self._browse_m_out,TEXT_3,height=30).pack(fill="x",padx=12,pady=(0,10))

        _divider(lp,"LIVE STATS",YELLOW)
        self._m_stats_lbl=ctk.CTkLabel(lp,text="0 parts  |  0 warnings",
                                        font=FM(11),text_color=TEXT_2)
        self._m_stats_lbl.pack(pady=(6,3),padx=14,anchor="w")
        self._m_warn_fr=ctk.CTkScrollableFrame(lp,fg_color=_blend("#FFFFFF",0.03,_CARD),
                                               corner_radius=8,height=90)
        self._m_warn_fr.pack(fill="x",padx=12,pady=(0,10))
        self._m_warn_labels=[]

        _divider(lp,"ACTIONS",YELLOW)
        self._btn_m_save=_pill_btn(lp,"①  Save Excel",self._save_manual_excel,YELLOW,height=36)
        self._btn_m_save.pack(fill="x",padx=12,pady=(8,5))
        self._btn_m_run=_run_btn(lp,"②  VALIDATE & GENERATE 3D",self._run_manual,YELLOW)
        self._btn_m_run.pack(fill="x",padx=12,pady=(0,8))

        sf=ctk.CTkFrame(lp,fg_color=_blend("#FFFFFF",0.04,_CARD),corner_radius=10)
        sf.pack(fill="x",padx=12,pady=(0,12))
        sr=ctk.CTkFrame(sf,fg_color="transparent"); sr.pack(fill="x",padx=10,pady=(8,4))
        self._m_status_dot=ctk.CTkLabel(sr,text="●",font=ctk.CTkFont(_FF,13),text_color=GREEN)
        self._m_status_dot.pack(side="left")
        self._m_status=ctk.CTkLabel(sr,text="  Ready",font=ctk.CTkFont(_FF,13,"bold"),text_color=GREEN)
        self._m_status.pack(side="left")
        self._m_prog=ctk.CTkProgressBar(sf,fg_color=_blend("#FFFFFF",0.08,_CARD),
                                         progress_color=YELLOW,height=6,corner_radius=3)
        self._m_prog.set(0); self._m_prog.pack(fill="x",padx=10,pady=(0,3))
        self._m_prog_lbl=ctk.CTkLabel(sf,text="",font=ctk.CTkFont(_FT,10),text_color=TEXT_3)
        self._m_prog_lbl.pack(pady=(0,8))

        # ── Right ──
        rp.columnconfigure(0,weight=1)
        sf2=ctk.CTkFrame(rp,fg_color="transparent")
        sf2.grid(row=0,column=0,sticky="ew",padx=10,pady=(10,0))
        self._manual_search_var=ctk.StringVar()
        self._manual_search_var.trace_add("write",self._filter_manual_rows)
        ctk.CTkEntry(sf2,textvariable=self._manual_search_var,
                     placeholder_text="🔍  Filter by Part No or Type…",height=34,
                     fg_color=_blend("#FFFFFF",0.06,_SURF),border_color=BORDER_M,
                     text_color=TEXT_1,font=ctk.CTkFont(_FT,12)
                     ).pack(side="left",fill="x",expand=True)

        tb=ctk.CTkFrame(rp,fg_color="transparent")
        tb.grid(row=1,column=0,sticky="ew",padx=10,pady=(10,2))
        ctk.CTkLabel(tb,text="ADD  ",font=ctk.CTkFont(_FF,11,"bold"),text_color=YELLOW).pack(side="left")
        for gt,col in [("Spur_Gear_3D",YELLOW),("Helical_Gear",ORANGE),("Ring_Gear_3D",TEAL),("Bevel_Gear",BLUE)]:
            _pill_btn(tb,f"+ {gt.split('_')[0]}",lambda t=gt: self._add_row(t),col,height=30
                      ).pack(side="left",padx=3)

        ai_sep=ctk.CTkFrame(rp,fg_color="transparent",height=22)
        ai_sep.grid(row=2,column=0,sticky="ew",padx=12,pady=(10,0))
        ctk.CTkFrame(ai_sep,fg_color=PURPLE,height=1).place(relx=0,rely=0.5,relwidth=1.0)
        ctk.CTkLabel(ai_sep,text="  CUSTOM BUILDERS (AI)  ",
                     font=ctk.CTkFont(_FT,9,"bold"),text_color=TEXT_3,
                     fg_color=GLASS_BASE).place(relx=0.03,rely=0.5,anchor="w")
        self._custom_btn_fr=ctk.CTkScrollableFrame(rp,fg_color="transparent",orientation="horizontal",height=52)
        self._custom_btn_fr.grid(row=3,column=0,sticky="ew",padx=10,pady=(2,6))

        self._rows_outer=ctk.CTkScrollableFrame(rp,fg_color=_blend("#FFFFFF",0.03,_SURF),
                                                corner_radius=10,border_width=1,border_color=BORDER_S)
        self._rows_outer.grid(row=4,column=0,sticky="nsew",padx=10,pady=(0,5))

        # ── Live build-status table (per-part progress) ─────────────────────
        tbl_hdr=ctk.CTkFrame(rp,fg_color="transparent")
        tbl_hdr.grid(row=5,column=0,sticky="ew",padx=10,pady=(4,0))
        ctk.CTkLabel(tbl_hdr,text="  ⚙  BUILD STATUS",
                     font=ctk.CTkFont(_FF,11,"bold"),text_color=YELLOW).pack(side="left")

        tbl_fr=ctk.CTkFrame(rp,fg_color=_blend("#FFFFFF",0.04,_SURF),
                            corner_radius=10,border_width=1,border_color=BORDER_S)
        tbl_fr.grid(row=6,column=0,sticky="ew",padx=10,pady=(2,4))
        self._m2_tbl=tk.Text(tbl_fr,height=5,
                             bg=_blend("#FFFFFF",0.04,_SURF),fg=TEXT_1,
                             font=FM(10),relief="flat",padx=10,pady=5,
                             state="disabled",cursor="arrow")
        for tag,col in [("ok",GREEN),("err",RED),("warn",ORANGE),("head",YELLOW)]:
            self._m2_tbl.tag_config(tag,foreground=col)
        sb2=ctk.CTkScrollbar(tbl_fr,command=self._m2_tbl.yview)
        self._m2_tbl.configure(yscrollcommand=sb2.set)
        sb2.pack(side="right",fill="y"); self._m2_tbl.pack(fill="both",expand=True)

        # ── Engine log ──────────────────────────────────────────────────────
        log_hdr=ctk.CTkFrame(rp,fg_color="transparent")
        log_hdr.grid(row=7,column=0,sticky="ew",padx=10,pady=(2,0))
        ctk.CTkLabel(log_hdr,text="  ENGINE LOG",
                     font=ctk.CTkFont(_FF,11,"bold"),text_color=YELLOW).pack(side="left")
        _pill_btn(log_hdr,"Clear",lambda: self._clear_log(self._log2_txt),
                  TEXT_3,height=24,width=55).pack(side="right",padx=4)

        log_fr,self._log2_txt=_log_widget(rp)
        log_fr.grid(row=8,column=0,sticky="nsew",padx=10,pady=(2,10))

        # Row weights: gear-row list gets most space, log gets remaining
        rp.rowconfigure(4,weight=3)
        rp.rowconfigure(8,weight=2)

        self._refresh_custom_buttons(); self._add_row("Spur_Gear_3D")

    def _add_row(self, gt="Spur_Gear_3D"):
        def _del(): self._gear_rows.remove(row); row.destroy(); self._update_manual_stats()
        row=GearRow(self._rows_outer,_del,self._update_manual_stats,self)
        row.v_type.set(gt); self._gear_rows.append(row); self._update_manual_stats()

    def _update_manual_stats(self):
        total=len(self._gear_rows); warns=[]
        for r in self._gear_rows:
            w=r._warn_lbl.cget("text")
            if w: warns.append(f"{r.v_pno.get()}: {w}")
        self._m_stats_lbl.configure(text=f"{total} parts  |  {len(warns)} warnings")
        for lbl in self._m_warn_labels:
            try: lbl.destroy()
            except Exception: pass
        self._m_warn_labels.clear()
        for w in warns[:8]:
            lbl=ctk.CTkLabel(self._m_warn_fr,text=w,font=ctk.CTkFont(_FT,10),text_color=ORANGE,anchor="w")
            lbl.pack(anchor="w",padx=5,pady=2); self._m_warn_labels.append(lbl)

    def _get_manual_parts(self): return [r.get_part() for r in self._gear_rows]

    def _browse_m_out(self):
        p=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],initialfile="manual_bom.xlsx")
        if p: self._m_out.set(p)

    def _save_manual_excel(self):
        parts=self._get_manual_parts()
        if not parts: messagebox.showwarning("Save","No parts defined."); return
        out=self._m_out.get(); os.makedirs(os.path.dirname(os.path.abspath(out)),exist_ok=True)
        try:
            write_bom_excel(parts,out); _append_log(self._log2_txt,f"✔ Excel saved: {out}","ok")
            messagebox.showinfo("Saved",f"Excel saved:\n{out}")
        except Exception as e: _append_log(self._log2_txt,f"✘ Save failed: {e}","err")

    def _run_manual(self):
        self._btn_m_run.configure(state="disabled")
        threading.Thread(target=self._t_run_manual,daemon=True).start()

    def _t_run_manual(self):
        parts=self._get_manual_parts()
        if not parts:
            self._q2.put(("log","✘ No parts defined")); self._q2.put(("btn2",None)); return
        out=self._m_out.get(); os.makedirs(os.path.dirname(os.path.abspath(out)),exist_ok=True)
        self._q2.put(("prog2",(0.05,"Saving Excel…")))
        try:
            write_bom_excel(parts,out); self._q2.put(("log",f"✔ Excel: {out}"))
        except Exception as e:
            self._q2.put(("log",f"✘ Excel write: {e}")); self._q2.put(("btn2",None)); return

        self._q2.put(("prog2",(0.12,"Validating…")))
        try:
            from validator_3d import Validator3D
            from autocad_engine_3d import AutoCAD3DGearEngine
            v=Validator3D(out,log_callback=lambda m: self._q2.put(("log",m)))
            v.run_checks(); vp=v.valid_parts
            if not vp:
                self._q2.put(("log","✘ Validation failed.")); self._q2.put(("btn2",None)); return

            n = len(vp)
            st = {p["Part_Number"]: "⏳ Queued" for p in vp}
            self._q2.put(("table2",(vp, dict(st))))
            self._q2.put(("prog2",(0.20,"Starting AutoCAD…")))
            time.sleep(0.05)   # let poller tick before heavy work starts

            eng = AutoCAD3DGearEngine(log_cb=lambda m: self._q2.put(("log",m)))

            # ── One part at a time so every update is visible ───────────────
            for i, part in enumerate(vp):
                pno = part["Part_Number"]
                st[pno] = "⚙ Building…"
                pct = 0.25 + 0.70 * (i / n)
                self._q2.put(("table2",(vp, dict(st))))
                self._q2.put(("prog2",(pct, f"Building {pno}  ({i+1}/{n})")))
                time.sleep(0.05)   # yield so poller can repaint before build starts

                try:
                    eng.generate_3d_batch([part])
                    st[pno] = "✔ Done"
                except Exception as part_err:
                    st[pno] = "✘ Error"
                    self._q2.put(("log", f"✘ {pno}: {part_err}"))

                self._q2.put(("table2",(vp, dict(st))))
                self._q2.put(("prog2",(0.25 + 0.70*((i+1)/n),
                                       f"Done {pno}  ({i+1}/{n})")))

            done  = sum(1 for s in st.values() if s=="✔ Done")
            error = sum(1 for s in st.values() if s=="✘ Error")
            col   = GREEN if error==0 else ORANGE
            self._q2.put(("prog2",(1.0,"Complete!")))
            self._q2.put(("status2",(f"✔ {done} built  {f'| ✘ {error} errors' if error else ''}",col)))
        except Exception as e:
            self._q2.put(("log",f"✘ {e}\n{traceback.format_exc()}"))
            self._q2.put(("status2",("● Error",RED)))
        self._q2.put(("btn2",None))

    # ═══════════════════════════════════════════════════════════════════════
    # TAB 3 — AI SHAPE CREATOR
    # ═══════════════════════════════════════════════════════════════════════
    def _build_ai_tab(self):
        lp, rp = self._make_two_col("ai", left_w=300, accent=PURPLE)
        # ── Left ──
        ctk.CTkFrame(lp,fg_color=PURPLE,height=3,corner_radius=2).pack(fill="x")
        ctk.CTkLabel(lp,text="🤖  AI SHAPE CREATOR",font=ctk.CTkFont(_FF,16,"bold"),text_color=PURPLE
                     ).pack(pady=(18,3),padx=14)
        ctk.CTkLabel(lp,text="Text-to-CAD via Gemini → CSG JSON → 3D",
                     font=ctk.CTkFont(_FT,11),text_color=TEXT_3,wraplength=260).pack(pady=(0,12))

        _divider(lp,"PART CONFIG",PURPLE)
        ctk.CTkLabel(lp,text="Part Name (becomes Custom_...):",font=ctk.CTkFont(_FT,11),text_color=TEXT_3
                     ).pack(anchor="w",padx=14,pady=(4,0))
        self._ai_part_name=ctk.StringVar(value="Sensor_Mount")
        _entry(lp,self._ai_part_name).pack(fill="x",padx=12,pady=(4,14))

        _divider(lp,"ACTIONS",PURPLE)
        self._btn_ai_gen=_run_btn(lp,"①  GENERATE JSON TEMPLATE",self._run_ai_generate,PURPLE)
        self._btn_ai_gen.pack(fill="x",padx=12,pady=(8,6))
        self._btn_ai_run=_run_btn(lp,"②  VALIDATE & BUILD 3D",self._run_ai_3d,_blend(PURPLE,0.75))
        self._btn_ai_run.pack(fill="x",padx=12,pady=(0,8))

        sf=ctk.CTkFrame(lp,fg_color=_blend("#FFFFFF",0.04,_CARD),corner_radius=10)
        sf.pack(fill="x",padx=12,pady=(0,12))
        sr=ctk.CTkFrame(sf,fg_color="transparent"); sr.pack(fill="x",padx=10,pady=(8,4))
        self._ai_status_dot=ctk.CTkLabel(sr,text="●",font=ctk.CTkFont(_FF,13),text_color=GREEN)
        self._ai_status_dot.pack(side="left")
        self._ai_status=ctk.CTkLabel(sr,text="  Ready",font=ctk.CTkFont(_FF,13,"bold"),text_color=GREEN)
        self._ai_status.pack(side="left")
        self._ai_prog=ctk.CTkProgressBar(sf,fg_color=_blend("#FFFFFF",0.08,_CARD),
                                          progress_color=PURPLE,height=6,corner_radius=3)
        self._ai_prog.set(0); self._ai_prog.pack(fill="x",padx=10,pady=(0,3))
        self._ai_prog_lbl=ctk.CTkLabel(sf,text="",font=ctk.CTkFont(_FT,10),text_color=TEXT_3)
        self._ai_prog_lbl.pack(pady=(0,8))

        # ── Right ──
        rp.rowconfigure(1,weight=2); rp.rowconfigure(3,weight=1); rp.columnconfigure(0,weight=1)
        ctk.CTkLabel(rp,text="  💬  DESCRIBE YOUR CUSTOM SHAPE",
                     font=ctk.CTkFont(_FF,12,"bold"),text_color=PURPLE
                     ).grid(row=0,column=0,sticky="w",padx=10,pady=(12,3))
        prompt_fr=ctk.CTkFrame(rp,fg_color=_blend("#FFFFFF",0.05,_SURF),
                               corner_radius=10,border_width=1,border_color=BORDER_S)
        prompt_fr.grid(row=1,column=0,sticky="nsew",padx=10,pady=(0,6))
        self._ai_prompt=tk.Text(prompt_fr,bg=_blend("#FFFFFF",0.05,_SURF),fg=TEXT_1,
                                font=(_FT,12),relief="flat",padx=14,pady=12,wrap="word",
                                insertbackground=PURPLE,selectbackground=_blend(PURPLE,0.25))
        self._ai_prompt.pack(fill="both",expand=True)
        self._ai_prompt.insert("end","Example: A rectangular base where length is P1, width is P2, "
                               "thickness is P3. Subtract a cylinder hole in the center with diameter P4.")
        ctk.CTkLabel(rp,text="  🤖  AI CSG LOG",font=ctk.CTkFont(_FF,11,"bold"),text_color=PURPLE
                     ).grid(row=2,column=0,sticky="w",padx=10,pady=(4,2))
        log_fr,self._log3_txt=_log_widget(rp)
        log_fr.grid(row=3,column=0,sticky="nsew",padx=10,pady=(0,10))
        _append_log(self._log3_txt,"SYSTEM  AI Mode — describe shape using P1, P2, P3, P4.","info")

    def _run_ai_generate(self):
        self._btn_ai_gen.configure(state="disabled")
        threading.Thread(target=self._t_ai_generate,daemon=True).start()

    def _t_ai_generate(self):
        desc=self._ai_prompt.get("1.0","end").strip()
        pname=self._ai_part_name.get().strip().replace(" ","_")
        key=os.environ.get("GEMINI_API_KEY","").strip()
        if not desc:  self._q3.put(("log","✘ Enter a description.")); self._q3.put(("btn_ai_gen",None)); return
        if not pname: self._q3.put(("log","✘ Enter a Part Name.")); self._q3.put(("btn_ai_gen",None)); return
        if not key:   self._q3.put(("log","✘ GEMINI_API_KEY env var missing.")); self._q3.put(("btn_ai_gen",None)); return
        self._q3.put(("prog3",(0.3,"Sending to Gemini…")))
        try:
            from genai_creator import generate_siraal_shape
            ok=generate_siraal_shape(part_name=pname,description=desc,api_key=key,
                                     model_name="gemini-2.5-flash",log_cb=lambda m: self._q3.put(("log",m)))
            if ok: self._q3.put(("status3",("✔ Template Ready",GREEN))); self._q3.put(("prog3",(1.0,"Ready!")))
            else:  self._q3.put(("prog3",(0,"Failed")))
        except ImportError: self._q3.put(("log","✘ genai_creator.py not found.")); self._q3.put(("prog3",(0,"Import Error")))
        except Exception as e: self._q3.put(("log",f"✘ {e}")); self._q3.put(("prog3",(0,"Error")))
        self._q3.put(("btn_ai_gen",None))

    def _run_ai_3d(self):
        pname=self._ai_part_name.get().strip().replace(" ","_")
        if pname.startswith("Custom_"): pname=pname[7:]
        tj=os.path.join("templates",f"Custom_{pname}.json")
        if not os.path.exists(tj): messagebox.showwarning("AI Mode",f"Run Step ① first."); return
        self._btn_ai_run.configure(state="disabled")
        threading.Thread(target=self._t_ai_3d,args=(pname,),daemon=True).start()

    def _t_ai_3d(self, pname):
        pythoncom.CoInitialize(); self._q3.put(("prog3",(0.1,"Creating preview BOM…")))
        pp=[{"Part_Number":f"AI-PREV-{pname.upper()[:10]}","Part_Type":f"Custom_{pname}",
             "Material":"Al-6061","Param_1":"100","Param_2":"100","Param_3":"20","Param_4":"15",
             "Quantity":1,"Priority":"High","Description":"AI Generated CSG Preview","Enabled":"YES"}]
        out=os.path.abspath("output_ai/ai_preview_bom.xlsx"); os.makedirs(os.path.dirname(out),exist_ok=True)
        try:
            write_bom_excel(pp,out)
            self._q3.put(("log",f"SYSTEM  BOM created for Custom_{pname}"))
        except Exception as e: self._q3.put(("log",f"✘ {e}")); self._q3.put(("btn_ai_run",None)); return
        self._q3.put(("prog3",(0.3,"Starting AutoCAD…")))
        try:
            from validator_3d import Validator3D
            from autocad_engine_3d import AutoCAD3DGearEngine
            v=Validator3D(out,log_callback=lambda m: self._q3.put(("log",m)))
            v.run_checks(); vp=v.valid_parts
            if not vp: self._q3.put(("log","✘ Validation failed.")); self._q3.put(("btn_ai_run",None)); return
            eng=AutoCAD3DGearEngine(log_cb=lambda m: self._q3.put(("log",m)))
            self._q3.put(("prog3",(0.6,"Building 3D…")))
            eng.generate_3d_batch(vp)
            self._q3.put(("prog3",(1.0,"Done!"))); self._q3.put(("status3",("✔ 3D Preview Built",GREEN)))
        except Exception as e:
            self._q3.put(("log",f"✘ {e}\n{traceback.format_exc()}")); self._q3.put(("status3",("● Error",RED)))
        self._q3.put(("btn_ai_run",None))

    # ═══════════════════════════════════════════════════════════════════════
    # TAB 4 — AI BOM COPILOT
    # ═══════════════════════════════════════════════════════════════════════
    def _build_copilot_tab(self):
        lp, rp = self._make_two_col("copilot", left_w=305, accent=BLUE)
        # ── Left ──
        ctk.CTkFrame(lp,fg_color=BLUE,height=3,corner_radius=2).pack(fill="x")
        ctk.CTkLabel(lp,text="🧠  AI BOM COPILOT",font=ctk.CTkFont(_FF,16,"bold"),text_color=BLUE
                     ).pack(pady=(18,3),padx=14)
        ctk.CTkLabel(lp,text="Edit mass Excel BOMs with natural language.",
                     font=ctk.CTkFont(_FT,11),text_color=TEXT_3,wraplength=275).pack(pady=(0,12))

        _divider(lp,"1.  TARGET BOM",BLUE)
        self._cp_file=ctk.StringVar(value="")
        _entry(lp,self._cp_file,placeholder="Select existing BOM…").pack(fill="x",padx=12,pady=(6,4))
        r=ctk.CTkFrame(lp,fg_color="transparent"); r.pack(fill="x",padx=12,pady=(0,10))
        _pill_btn(r,"Browse",self._browse_cp,BLUE,height=30).pack(side="left")

        _divider(lp,"2.  INSTRUCTIONS",BLUE)
        ctk.CTkLabel(lp,text="What should the AI change?",font=ctk.CTkFont(_FT,11),text_color=TEXT_3
                     ).pack(anchor="w",padx=14,pady=(4,0))
        self._cp_prompt=tk.Text(lp,bg=_blend("#FFFFFF",0.05,_CARD),fg=TEXT_1,
                                font=(_FT,11),relief="flat",padx=10,pady=10,
                                wrap="word",insertbackground=BLUE,height=7)
        self._cp_prompt.pack(fill="x",padx=12,pady=(4,12))
        self._cp_prompt.insert("end","Example: Change the material of all Spur Gears to Al-6061 and increase Face Width by 5mm.")

        _divider(lp,"ACTIONS",BLUE)
        self._btn_cp_preview=_run_btn(lp,"①  PREVIEW AI CHANGES",self._run_cp_preview,BLUE,height=38)
        self._btn_cp_preview.pack(fill="x",padx=12,pady=(8,5))
        self._btn_cp_apply=_pill_btn(lp,"②  APPROVE & SAVE EXCEL",self._run_cp_apply,GREEN,height=36)
        self._btn_cp_apply.pack(fill="x",padx=12,pady=(0,8)); self._btn_cp_apply.configure(state="disabled")

        sf=ctk.CTkFrame(lp,fg_color=_blend("#FFFFFF",0.04,_CARD),corner_radius=10)
        sf.pack(fill="x",padx=12,pady=(0,12))
        sr=ctk.CTkFrame(sf,fg_color="transparent"); sr.pack(fill="x",padx=10,pady=(8,4))
        self._cp_status_dot=ctk.CTkLabel(sr,text="●",font=ctk.CTkFont(_FF,13),text_color=TEXT_3)
        self._cp_status_dot.pack(side="left")
        self._cp_status=ctk.CTkLabel(sr,text="  Waiting",font=ctk.CTkFont(_FF,13,"bold"),text_color=TEXT_3)
        self._cp_status.pack(side="left")
        self._cp_prog=ctk.CTkProgressBar(sf,fg_color=_blend("#FFFFFF",0.08,_CARD),
                                          progress_color=BLUE,height=6,corner_radius=3)
        self._cp_prog.set(0); self._cp_prog.pack(fill="x",padx=10,pady=(0,3))
        self._cp_prog_lbl=ctk.CTkLabel(sf,text="",font=ctk.CTkFont(_FT,10),text_color=TEXT_3)
        self._cp_prog_lbl.pack(pady=(0,8))

        # ── Right ──
        rp.rowconfigure(1,weight=2); rp.rowconfigure(3,weight=1); rp.columnconfigure(0,weight=1)
        ctk.CTkLabel(rp,text="  📊  DATA DIFF VIEWER",font=ctk.CTkFont(_FF,12,"bold"),text_color=BLUE
                     ).grid(row=0,column=0,sticky="w",padx=10,pady=(12,3))
        diff_fr=ctk.CTkFrame(rp,fg_color=_blend("#FFFFFF",0.05,_SURF),
                             corner_radius=10,border_width=1,border_color=BORDER_S)
        diff_fr.grid(row=1,column=0,sticky="nsew",padx=10,pady=(0,6))
        self._cp_diff_txt=tk.Text(diff_fr,bg=_blend("#FFFFFF",0.05,_SURF),fg=GREEN,
                                  font=FM(10),relief="flat",padx=10,pady=8,state="disabled",wrap="none")
        sb_y=ctk.CTkScrollbar(diff_fr,command=self._cp_diff_txt.yview)
        sb_x=ctk.CTkScrollbar(diff_fr,command=self._cp_diff_txt.xview,orientation="horizontal")
        self._cp_diff_txt.configure(yscrollcommand=sb_y.set,xscrollcommand=sb_x.set)
        sb_y.pack(side="right",fill="y"); sb_x.pack(side="bottom",fill="x"); self._cp_diff_txt.pack(fill="both",expand=True)

        ctk.CTkLabel(rp,text="  🧠  COPILOT REASONING LOG",font=ctk.CTkFont(_FF,11,"bold"),text_color=BLUE
                     ).grid(row=2,column=0,sticky="w",padx=10,pady=(4,2))
        log_fr,self._log4_txt=_log_widget(rp)
        log_fr.grid(row=3,column=0,sticky="nsew",padx=10,pady=(0,10))
        _append_log(self._log4_txt,"SYSTEM  AI Copilot initialized. Awaiting commands.","info")

    def _browse_cp(self):
        p=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xls"),("All","*.*")])
        if p: self._cp_file.set(p)

    def _run_cp_preview(self):
        self._btn_cp_preview.configure(state="disabled"); self._btn_cp_apply.configure(state="disabled")
        threading.Thread(target=self._t_cp_preview,daemon=True).start()

    def _t_cp_preview(self):
        ep=self._cp_file.get(); prompt=self._cp_prompt.get("1.0","end").strip()
        key=os.environ.get("GEMINI_API_KEY","").strip()
        if not ep or not os.path.exists(ep):
            self._q4.put(("log","✘ Select a valid Excel file first.")); self._q4.put(("preview_fail",None)); return
        self._q4.put(("prog4",(0.2,"Reading & Analyzing…"))); self._q4.put(("status4",("● Processing",BLUE)))
        def tl(m): self._q4.put(("log",m))
        try:
            from ai_bom_copilot import preview_bom_edits
            ok,nd,dt=preview_bom_edits(ep,prompt,key,tl)
            if ok:
                self._q4.put(("diff",dt)); self._q4.put(("preview_success",nd))
                self._q4.put(("prog4",(1.0,"Preview Ready!"))); self._q4.put(("status4",("✔ Ready for Approval",GREEN)))
            else: self._q4.put(("preview_fail",None)); self._q4.put(("status4",("● Error",RED)))
        except ImportError: self._q4.put(("log","✘ ai_bom_copilot.py not found.")); self._q4.put(("preview_fail",None))
        except Exception as e: self._q4.put(("log",f"✘ {e}")); self._q4.put(("preview_fail",None))

    def _run_cp_apply(self):
        if not self._pending_copilot_data: return
        out=self._cp_file.get(); prompt=self._cp_prompt.get("1.0","end").strip()
        self._btn_cp_apply.configure(state="disabled")
        threading.Thread(target=self._t_cp_apply,args=(out,prompt),daemon=True).start()

    def _t_cp_apply(self, out, prompt):
        try:
            self._q4.put(("log","SYSTEM  Committing changes to Excel…"))
            from ai_bom_copilot import commit_bom_edits
            ok,msg=commit_bom_edits(excel_path=out,new_dicts=self._pending_copilot_data,
                                    log_cb=lambda m: self._q4.put(("log",m)),
                                    author="AI Copilot via GUI",prompt=prompt)
            if ok: self._q4.put(("apply_success",msg))
            else:  self._q4.put(("apply_fail",msg))
        except Exception as e: self._q4.put(("log",f"✘ {e}")); self._q4.put(("apply_fail",str(e)))

    # ── Watcher bar ────────────────────────────────────────────────────────
    def _build_watcher_bar(self):
        ctk.CTkFrame(self,fg_color=BORDER_S,height=1,corner_radius=0).pack(fill="x",side="bottom")
        bar=ctk.CTkFrame(self,fg_color=_blend("#FFFFFF",0.04,_SURF),corner_radius=0,height=44)
        bar.pack(fill="x",side="bottom"); bar.pack_propagate(False)
        ctk.CTkLabel(bar,text="⚡  REALTIME WATCHER",font=ctk.CTkFont(_FF,11,"bold"),text_color=YELLOW
                     ).pack(side="left",padx=(14,8),pady=8)
        self._watch_path_var=ctk.StringVar(value="")
        ctk.CTkEntry(bar,textvariable=self._watch_path_var,width=330,height=28,
                     fg_color=_blend("#FFFFFF",0.06,_SURF),border_color=BORDER_M,
                     text_color=TEXT_2,font=ctk.CTkFont(_FT,10),
                     placeholder_text="Select Excel file to watch…").pack(side="left",padx=(0,6))
        ctk.CTkButton(bar,text="Browse",width=70,height=28,
                      fg_color=_blend("#FFFFFF",0.07,_SURF),hover_color=_blend("#FFFFFF",0.12,_SURF),
                      text_color=TEXT_2,font=ctk.CTkFont(_FT,9),
                      command=self._browse_watch).pack(side="left",padx=(0,6))
        self._watch_toggle=ctk.CTkButton(bar,text="▶  START",width=100,height=28,
                                          fg_color=_blend(TEAL,0.15,_SURF),hover_color=_blend(TEAL,0.25,_SURF),
                                          border_color=TEAL,border_width=1,text_color=TEAL,
                                          font=ctk.CTkFont(_FF,10,"bold"),command=self._toggle_watch)
        self._watch_toggle.pack(side="left",padx=(0,10))
        self._watch_indicator=ctk.CTkLabel(bar,text="● IDLE",font=ctk.CTkFont(_FF,10,"bold"),text_color=TEXT_3)
        self._watch_indicator.pack(side="left")
        self._watch_diff_fr=ctk.CTkFrame(bar,fg_color=_blend("#FFFFFF",0.04,_SURF),corner_radius=6,height=28)
        self._watch_diff_fr.pack(side="left",fill="x",expand=True,padx=(12,12),pady=8)
        self._watch_diff_lbl=ctk.CTkLabel(self._watch_diff_fr,text="No changes detected",
                                           font=FM(9),text_color=TEXT_3,anchor="w")
        self._watch_diff_lbl.pack(fill="x",padx=8)

    def _browse_watch(self):
        p=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xls"),("All","*.*")])
        if p: self._watch_path_var.set(p)

    def _toggle_watch(self):
        if self._watch_active:
            self._watcher.stop(); self._watch_active=False
            self._watch_toggle.configure(text="▶  START",fg_color=_blend(TEAL,0.15,_SURF),border_color=TEAL,text_color=TEAL)
            self._watch_indicator.configure(text="● IDLE",text_color=TEXT_3)
        else:
            p=self._watch_path_var.get()
            if not p or not os.path.exists(p): messagebox.showwarning("Watch","Select a valid Excel file first."); return
            self._watcher.start(p); self._watch_active=True
            self._watch_toggle.configure(text="⏹  STOP",fg_color=_blend(RED,0.15,_SURF),border_color=RED,text_color=RED)
            self._watch_indicator.configure(text=f"● WATCHING  {os.path.basename(p)}",text_color=GREEN)

    def _on_excel_change(self, diffs):
        preview=" │ ".join(diffs[:3])
        if len(diffs)>3: preview+=f"  +{len(diffs)-3} more"
        def _ui():
            self._watch_diff_lbl.configure(
                text=f"⚡ {datetime.datetime.now().strftime('%H:%M:%S')}  {preview}",text_color=YELLOW)
            for log in (self._log1_txt,self._log2_txt,self._log3_txt,self._log4_txt):
                _append_log(log,f"⚡ Excel changed: {preview}","warn")
            messagebox.showinfo("Watched BOM Modified",
                                "The watched Excel file was modified externally.\nRe-Run validation to apply changes.")
        self.after(0,_ui)

    # ── Global Cost ────────────────────────────────────────────────────────
    def _run_global_cost(self):
        self._btn_global_cost.configure(state="disabled")
        threading.Thread(target=self._t_run_global_cost,daemon=True).start()

    def _t_run_global_cost(self):
        q_map={"excel":(self._q1,"prog1"),"manual":(self._q2,"prog2"),
               "ai":(self._q3,"prog3"),"copilot":(self._q4,"prog4")}
        q,pk=q_map.get(self._active_tab,(self._q1,"prog1"))
        try:
            parts=[]
            if self._active_tab=="manual":
                parts=self._get_manual_parts()
                if not parts: q.put(("log","✘ No parts defined.")); return
            else:
                path=""
                if self._active_tab=="excel":    path=self._e1_file.get()
                elif self._active_tab=="copilot": path=self._cp_file.get()
                if not path or not os.path.exists(path):
                    path=filedialog.askopenfilename(title="Select BOM",filetypes=[("Excel","*.xlsx *.xls")])
                    if not path: return
                q.put(("log",f"SYSTEM  ESG report for {os.path.basename(path)}…"))
                from validator_3d import Validator3D
                v=Validator3D(path,log_callback=lambda m: None)
                v.run_checks(); parts=v.valid_parts
            if not parts: q.put(("log","✘ No valid parts.")); return
            q.put((pk,(0.3,"Analyzing costs & carbon footprint…")))
            from cost_engine import CostEngine
            engine=CostEngine(metal_api_key=os.environ.get("METALPRICE_API_KEY",""),
                              gemini_api_key=os.environ.get("GEMINI_API_KEY","").strip())
            engine.fetch_live_metal_prices()
            out=os.path.abspath(f"output_reports/Siraal_ESG_{int(time.time())}.pdf")
            ok=engine.export_pdf_report(parts,out)
            if ok:
                q.put(("log",f"✔ PDF: {out}")); q.put((pk,(1.0,"Report Ready!")))
                try: os.startfile(out)
                except Exception as e: q.put(("log",f"Auto-open: {e}"))
            else: q.put(("log","✘ PDF failed (is fpdf2 installed?)"))
        except Exception as e: q.put(("log",f"✘ {e}\n{traceback.format_exc()}"))
        finally: self.after(0,lambda: self._btn_global_cost.configure(state="normal"))

    # ── Queue Pollers ──────────────────────────────────────────────────────
    def _poll_q1(self):
        try:
            for _ in range(10):   # max 10 items per tick so UI can repaint between ticks
                k,d=self._q1.get_nowait()
                if   k=="log":    _append_log(self._log1_txt,str(d))
                elif k=="table1": self._update_tbl(self._e1_tbl,d[0],d[1])
                elif k=="stats":  self._e1_stats.configure(text=str(d))
                elif k=="prog1":
                    self._e1_prog.set(d[0]); self._e1_prog_lbl.configure(text=d[1])
                    self.update_idletasks()   # force immediate repaint of progress bar
                elif k=="status1":
                    self._e1_status.configure(text=f"  {d[0]}",text_color=d[1])
                    self._e1_status_dot.configure(text_color=d[1])
                elif k=="btn1":   self._btn_e1.configure(state="normal")
        except queue.Empty: pass
        except Exception as e: print(f"Q1: {e}")
        finally: self.after(80,self._poll_q1)

    def _poll_q2(self):
        try:
            for _ in range(10):
                k,d=self._q2.get_nowait()
                if   k=="log":    _append_log(self._log2_txt,str(d))
                elif k=="table2": self._update_tbl(self._m2_tbl,d[0],d[1])
                elif k=="prog2":
                    self._m_prog.set(d[0]); self._m_prog_lbl.configure(text=d[1])
                    self.update_idletasks()
                elif k=="status2":
                    self._m_status.configure(text=f"  {d[0]}",text_color=d[1])
                    self._m_status_dot.configure(text_color=d[1])
                elif k=="btn2":   self._btn_m_run.configure(state="normal")
        except queue.Empty: pass
        except Exception as e: print(f"Q2: {e}")
        finally: self.after(80,self._poll_q2)

    def _poll_q3(self):
        try:
            for _ in range(10):
                k,d=self._q3.get_nowait()
                if   k=="log":        _append_log(self._log3_txt,str(d))
                elif k=="prog3":
                    self._ai_prog.set(d[0]); self._ai_prog_lbl.configure(text=d[1])
                    self.update_idletasks()
                elif k=="status3":
                    self._ai_status.configure(text=f"  {d[0]}",text_color=d[1])
                    self._ai_status_dot.configure(text_color=d[1])
                elif k=="btn_ai_gen": self._btn_ai_gen.configure(state="normal")
                elif k=="btn_ai_run": self._btn_ai_run.configure(state="normal")
        except queue.Empty: pass
        except Exception as e: print(f"Q3: {e}")
        finally: self.after(80,self._poll_q3)

    def _poll_q4(self):
        try:
            for _ in range(10):
                k,d=self._q4.get_nowait()
                if   k=="log":    _append_log(self._log4_txt,str(d))
                elif k=="prog4":
                    self._cp_prog.set(d[0]); self._cp_prog_lbl.configure(text=d[1])
                    self.update_idletasks()
                elif k=="status4":
                    self._cp_status.configure(text=f"  {d[0]}",text_color=d[1])
                    self._cp_status_dot.configure(text_color=d[1])
                elif k=="diff":
                    self._cp_diff_txt.configure(state="normal")
                    self._cp_diff_txt.delete("1.0","end")
                    self._cp_diff_txt.insert("end",d)
                    self._cp_diff_txt.configure(state="disabled")
                elif k=="preview_success":
                    self._pending_copilot_data=d
                    self._btn_cp_preview.configure(state="normal")
                    self._btn_cp_apply.configure(state="normal")
                elif k=="preview_fail":
                    self._btn_cp_preview.configure(state="normal")
                    self._btn_cp_apply.configure(state="disabled")
                elif k=="apply_success":
                    messagebox.showinfo("Copilot",d)
                    self._cp_status.configure(text="  Excel Saved",text_color=GREEN)
                    self._btn_cp_preview.configure(state="normal")
                elif k=="apply_fail":
                    self._btn_cp_apply.configure(state="normal")
                    self._btn_cp_preview.configure(state="normal")
        except queue.Empty: pass
        except Exception as e: print(f"Q4: {e}")
        finally: self.after(80,self._poll_q4)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
if __name__ == "__main__":
    app = SiraalGUI()
    app.mainloop()