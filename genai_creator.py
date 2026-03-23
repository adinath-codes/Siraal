"""
genai_creator.py  —  Siraal AI Shape Compiler  v2.0
====================================================
Converts natural language part descriptions into validated Siraal JSON
CSG recipes, saves them to templates/, and optionally adds the part
directly to the BOM Excel for immediate engine pickup.

PUBLIC API
──────────
  # Generate a template (does NOT touch Excel)
  ok = generate_siraal_shape("Turbine_Disc", "...", api_key="...", log_cb=print)

  # Generate + add to BOM + ready to run engine
  ok = generate_and_queue(
      part_name    = "Turbine_Disc",
      description  = "...",
      material     = "Steel-4140",
      p1=200, p2=30, p3=60, p4=40,
      excel_path   = "excels/demo_gears_3d.xlsx",
      api_key      = "YOUR_KEY",
      log_cb       = print,
  )
"""

from __future__ import annotations

import json
import math
import os
import re
from typing import Callable, Optional

try:
    from google import genai
    from google.genai import types as gtypes
    _GENAI_OK = True
except ImportError:
    _GENAI_OK = False

# ══════════════════════════════════════════════════════════════════════════════
#  SCHEMA REFERENCE  (embedded in every Gemini prompt)
# ══════════════════════════════════════════════════════════════════════════════

_SCHEMA = """
╔══════════════════════════════════════════════════════════════════════════════╗
║  SIRAAL JSON CSG SCHEMA  v3.0  —  Enterprise Grade                         ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  PARAMETERS                                                                 ║
║  P1-P4 come from the BOM Excel row.                                        ║
║  Extra computed values: use SET_VAR steps to define V1-V9.                 ║
║  All expressions may use:                                                   ║
║    P1 P2 P3 P4  V1..V9                                                     ║
║    math  pi  abs  min  max  sqrt  sin  cos  tan  asin  acos  atan  atan2   ║
║    hypot  ceil  floor  log  log10  exp  radians  degrees  pow  round        ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  TOP-LEVEL STRUCTURE                                                        ║
║  {                                                                          ║
║    "Part_Name":   "string",                                                 ║
║    "Description": "string",                                                 ║
║    "Steps": [ ...step objects... ]                                          ║
║  }                                                                          ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  EVERY GEOMETRY STEP HAS:                                                   ║
║    "action"     : BASE | ADD | SUBTRACT | INTERSECT                        ║
║    "shape"      : see SHAPES below                                          ║
║    "x_offset"   : float/expr — X offset from part centre                   ║
║    "y_offset"   : float/expr — Y offset from part centre                   ║
║    "z"          : float/expr — Z base position                              ║
║    "rotate_axis": "X" | "Y" | "Z"   (applied BEFORE Boolean)               ║
║    "rotate_deg" : float/expr                                                ║
║    "overlap"    : float/expr — extend shape by N mm BELOW z for guaranteed  ║
║                   Boolean fusion. Use 3-5mm whenever adding onto existing   ║
║                   geometry. Prevents floating disconnected parts.           ║
║    "origin"     : "centre" (default) | "corner"                            ║
║                   "corner" → box/pipe starts AT (curr_x, curr_y)           ║
║                   "centre" → box/pipe CENTRED on (curr_x, curr_y)          ║
║                   CRITICAL RULE: extrude_profile always starts at corner.  ║
║                   Use origin="corner" on box/pipe to match its X alignment.║
╠══════════════════════════════════════════════════════════════════════════════╣
║  ACTIONS — GEOMETRY                                                         ║
║  BASE        First solid. MUST be step 0. One per recipe.                  ║
║  ADD         Boolean union onto BASE.                                       ║
║  SUBTRACT    Boolean subtract from BASE.  Use z="-5", height="dim+10".     ║
║  INTERSECT   Boolean intersection — keeps only the overlapping volume.      ║
║                Fields: same as any geometry step (shape + dims).           ║
║                Use for: dome caps, clipped solids, lens shapes.            ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  ACTIONS — TRANSFORM                                                        ║
║  SCALE       Uniformly scale the current BASE.                             ║
║                Fields: factor (float/expr, e.g. "P1/P2")                  ║
║  SLICE       Cut BASE with a plane, discard one side.                      ║
║                Fields: plane_normal ("X"|"Y"|"Z"), plane_z (float/expr),   ║
║                        keep ("+" for above/right, "-" for below/left)      ║
║                        slab_size (optional, default P1*4)                  ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  ACTIONS — SWEEP                                                            ║
║  SWEEP       Sweep a 2D profile along a 3D spline path.                    ║
║                Fields:                                                      ║
║                  profile_points: flat [x,y,x,y,...] 2D closed profile      ║
║                  path_points:    flat [x,y,z, x,y,z,...] 3D path (≥2 pts) ║
║                  z:              elevation of profile plane                 ║
║                  boolean:        "ADD" | "SUBTRACT" (default "ADD")        ║
║                Use for: cooling channels, pipe bends, cam followers        ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  ACTIONS — VARIABLE                                                         ║
║  SET_VAR     Compute and store an intermediate value.                       ║
║                Fields: var ("V1"-"V9"), expr (expression string)           ║
║                Place BEFORE the steps that use the variable.               ║
║                Example: {"action":"SET_VAR","var":"V1","expr":"P1*0.707"}  ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  ACTIONS — PATTERN (all immediately follow the ADD/SUBTRACT template step) ║
║  PATTERN_CIRCULAR   Fields: count, total_angle, [center_x, center_y]      ║
║  PATTERN_LINEAR     Fields: count, dx, dy, dz                              ║
║  PATTERN_RADIAL     Non-uniform angles. Fields: angles ([deg1,deg2,...])   ║
║                     Use for: crankshaft throws, asymmetric bolt patterns   ║
║  ARRAY_GRID         Rectangular grid. Fields: rows, cols, dx, dy,          ║
║                     [start_x, start_y]                                     ║
║                     Use for: heatsink fins, bolt hole grids, cooling arrays ║
║  HELIX_ARRAY        Copies rising along Z+rotation. Fields: count, dz,    ║
║                     da_deg, [center_x, center_y]                          ║
║                     Use for: cooling fins along a curved surface, threads  ║
║  MIRROR             Fields: plane = "XZ" | "YZ" | "XY"                    ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  ACTIONS — NOTE ONLY (applied manually in AutoCAD after build)             ║
║  FILLET      Fields: radius                                                 ║
║  CHAMFER     Fields: distance                                               ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  SHAPES & REQUIRED FIELDS                                                   ║
║  cylinder       radius, height                                              ║
║  box            length, width, height                                       ║
║  sphere         radius                                                      ║
║  cone           radius, height                                              ║
║  torus          major_radius, minor_radius                                  ║
║  frustum        radius_bottom, radius_top, height                           ║
║  pipe           outer_radius, inner_radius, height          ← NEW          ║
║  polygon_prism  sides, radius, height                       ← NEW          ║
║  ellipsoid      rx, ry, rz                                  ← NEW          ║
║  spring         coil_radius, wire_radius, pitch, turns      ← NEW          ║
║  extrude_profile  points [x,y,...], height, [taper_angle]                  ║
║  revolve          profile_points [x,y,...], axis (X/Y/Z), [degrees]        ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  GOLDEN RULES                                                               ║
║  1.  Step 0 MUST be action=BASE. One BASE per recipe.                       ║
║  2.  BASE must enclose the largest volume.                                  ║
║  3.  SUBTRACT tools: z="-5", height="dim+10" for through-cuts.             ║
║  4.  PATTERN rule: template step then PATTERN_ immediately after.          ║
║      The engine builds ALL N instances — do NOT pre-cut then pattern rest. ║
║  5.  SET_VAR steps go FIRST, before any geometry step that uses the var.   ║
║  6.  INTERSECT, SLICE, SCALE, SWEEP go BEFORE patterns/mirrors.            ║
║  7.  MIRROR, PATTERN, FILLET, CHAMFER go LAST.                             ║
║  8.  Expressions: "P1/2", "V1*cos(radians(P4))", "sqrt(P1**2+P2**2)"       ║
║  9.  revolve: profile_points are 2D (X=radial offset, Y=axial position)    ║
║  10. Keep recipes to 8-16 steps. More steps = longer build time.           ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

# ══════════════════════════════════════════════════════════════════════════════
#  FEW-SHOT EXAMPLES  (shown to Gemini so it learns the exact pattern)
# ══════════════════════════════════════════════════════════════════════════════

_EXAMPLES = """
════════════════════════════════════════════════════════
EXAMPLE 1 — Turbine Disc  (SET_VAR + fir-tree slots + PATTERN_CIRCULAR)
P1=OD_mm  P2=bore_mm  P3=thickness_mm  P4=slot_count
════════════════════════════════════════════════════════
{
  "Part_Name": "Turbine_Disc",
  "Steps": [
    { "action":"SET_VAR", "var":"V1", "expr":"P1*0.035",
      "note":"slot half-width" },
    { "action":"SET_VAR", "var":"V2", "expr":"P1*0.06",
      "note":"slot radial depth" },
    { "action":"BASE", "shape":"cylinder",
      "radius":"P1/2", "height":"P3", "z":"0" },
    { "action":"SUBTRACT", "shape":"cylinder",
      "radius":"P2/2", "height":"P3+10", "z":"-5" },
    { "action":"SUBTRACT", "shape":"extrude_profile",
      "height":"P3+10", "z":"-5",
      "x_offset":"P1/2 - V2/2",
      "points":[
        "-V1*1.5","0",  "-V1","V2*0.28",  "-V1*0.7","V2*0.37",
        "-V1*0.7","V2*0.64",  "-V1","V2*0.73",  "-V1","V2",
        "V1","V2",  "V1","V2*0.73",  "V1*0.7","V2*0.64",
        "V1*0.7","V2*0.37",  "V1","V2*0.28",  "V1*1.5","0"
      ],
      "note":"3-lobe fir-tree slot profile — centred at rim" },
    { "action":"PATTERN_CIRCULAR", "count":"P4", "total_angle":360 },
    { "action":"SUBTRACT", "shape":"cylinder",
      "radius":"P1*0.025", "height":"P3+10", "z":"-5",
      "x_offset":"P1*0.30" },
    { "action":"PATTERN_CIRCULAR", "count":8, "total_angle":360 },
    { "action":"FILLET", "radius":"3" }
  ]
}

════════════════════════════════════════════════════════
EXAMPLE 2 — Crankshaft Journal (PATTERN_RADIAL + SET_VAR + revolve)
P1=journal_dia  P2=crankpin_dia  P3=stroke  P4=n_cylinders
Builds a 4-cylinder crankshaft skeleton
════════════════════════════════════════════════════════
{
  "Part_Name": "Crankshaft_4cyl",
  "Steps": [
    { "action":"SET_VAR", "var":"V1", "expr":"P3/2",
      "note":"throw = half stroke" },
    { "action":"SET_VAR", "var":"V2", "expr":"P3*0.18",
      "note":"web thickness" },
    { "action":"SET_VAR", "var":"V3", "expr":"P3*4 + P3*0.6",
      "note":"total shaft length" },
    { "action":"BASE", "shape":"cylinder",
      "radius":"P1/2", "height":"V3", "z":"0",
      "note":"Main journal spine" },
    { "action":"ADD", "shape":"cylinder",
      "radius":"P1*0.65", "height":"V2",
      "z":"P3*0.4",
      "note":"One counterweight web — PATTERN_LINEAR replicates" },
    { "action":"PATTERN_LINEAR",
      "count":"P4", "dx":"0", "dy":"0", "dz":"P3*1.0" },
    { "action":"ADD", "shape":"cylinder",
      "radius":"P2/2", "height":"P3*0.7",
      "x_offset":"V1", "z":"P3*0.15",
      "note":"One crankpin throw at 0° — PATTERN_RADIAL gives 90/180/270" },
    { "action":"PATTERN_RADIAL",
      "angles":[0, 90, 270, 180],
      "note":"V8-style firing order offsets" },
    { "action":"SUBTRACT", "shape":"cylinder",
      "radius":"P1*0.25", "height":"V3+10", "z":"-5",
      "note":"Central oil drilling bore" },
    { "action":"FILLET", "radius":"5",
      "note":"Stress relief at all journal-web transitions" }
  ]
}

════════════════════════════════════════════════════════
EXAMPLE 3 — Heat Exchanger End Cap (ARRAY_GRID + pipe + revolve)
P1=shell_OD  P2=tube_OD  P3=cap_thickness  P4=tube_rows
════════════════════════════════════════════════════════
{
  "Part_Name": "HX_Tubesheet",
  "Steps": [
    { "action":"SET_VAR", "var":"V1", "expr":"P2*1.25",
      "note":"tube pitch = 1.25 × OD" },
    { "action":"BASE", "shape":"cylinder",
      "radius":"P1/2", "height":"P3", "z":"0",
      "note":"Circular tubesheet disc" },
    { "action":"ADD", "shape":"pipe",
      "outer_radius":"P1/2", "inner_radius":"P1/2 - P3*0.4",
      "height":"P3*0.6", "z":"P3",
      "note":"Rim flange" },
    { "action":"SUBTRACT", "shape":"cylinder",
      "radius":"P2/2", "height":"P3+10", "z":"-5",
      "x_offset":"V1",
      "note":"One tube bore template" },
    { "action":"ARRAY_GRID",
      "rows":"P4", "cols":"P4",
      "dx":"V1", "dy":"V1",
      "start_x":"-(P4-1)*V1/2",
      "start_y":"-(P4-1)*V1/2",
      "note":"Square pitch tube array" },
    { "action":"SUBTRACT", "shape":"cylinder",
      "radius":"P1*0.08", "height":"P3+10", "z":"-5",
      "x_offset":"P1*0.38",
      "note":"One bolt hole" },
    { "action":"PATTERN_CIRCULAR", "count":12, "total_angle":360 }
  ]
}

════════════════════════════════════════════════════════
EXAMPLE 4 — Helical Compression Spring (spring shape)
P1=coil_OD  P2=wire_dia  P3=free_length  P4=n_coils
════════════════════════════════════════════════════════
{
  "Part_Name": "Compression_Spring",
  "Steps": [
    { "action":"BASE", "shape":"spring",
      "coil_radius":"P1/2 - P2/2",
      "wire_radius":"P2/2",
      "pitch":"P3/P4",
      "turns":"P4",
      "z":"0",
      "note":"Full helical spring" },
    { "action":"SLICE", "plane_normal":"Z",
      "plane_z":"P3*0.04", "keep":"+",
      "note":"Grind bottom end flat" },
    { "action":"SLICE", "plane_normal":"Z",
      "plane_z":"P3*0.96", "keep":"-",
      "note":"Grind top end flat" }
  ]
}

════════════════════════════════════════════════════════
EXAMPLE 5 — Axisymmetric Nozzle with Dome Cap (revolve + INTERSECT + SLICE)
P1=inlet_OD  P2=throat_OD  P3=exit_OD  P4=total_length
════════════════════════════════════════════════════════
{
  "Part_Name": "Nozzle_Converging_Diverging",
  "Steps": [
    { "action":"SET_VAR","var":"V1","expr":"P4*0.35",
      "note":"converging section length" },
    { "action":"SET_VAR","var":"V2","expr":"P4*0.65",
      "note":"diverging section length" },
    { "action":"BASE", "shape":"revolve",
      "axis":"Z", "degrees":360,
      "profile_points":[
        "P1/2","0",
        "P1/2","P4*0.06",
        "P2/2","V1",
        "P2/2","V1+P4*0.04",
        "P3/2","P4",
        "P3/2 - P4*0.03","P4",
        "P2/2 - P4*0.02","V1+P4*0.04",
        "P2/2 - P4*0.02","V1",
        "P1/2 - P4*0.03","P4*0.06",
        "P1/2 - P4*0.03","0"
      ],
      "note":"Convergent-divergent nozzle wall with wall thickness" }
  ]
}

════════════════════════════════════════════════════════
EXAMPLE 6 — Turbine Stator Vane Carrier (HELIX_ARRAY + polygon_prism)
P1=carrier_OD  P2=vane_chord  P3=carrier_height  P4=n_vanes
════════════════════════════════════════════════════════
{
  "Part_Name": "Stator_Carrier",
  "Steps": [
    { "action":"SET_VAR","var":"V1","expr":"P1*0.08",
      "note":"carrier wall thickness" },
    { "action":"BASE", "shape":"pipe",
      "outer_radius":"P1/2",
      "inner_radius":"P1/2 - V1",
      "height":"P3", "z":"0" },
    { "action":"ADD", "shape":"extrude_profile",
      "height":"P3", "z":"0",
      "x_offset":"P1/2 - V1*0.5",
      "points":[
        "0","-P2*0.06",
        "P2*0.92","0",
        "P2","P2*0.12",
        "P2*0.15","P2*0.22",
        "0","P2*0.06"
      ],
      "note":"One aerofoil vane profile — HELIX_ARRAY staggers them axially" },
    { "action":"HELIX_ARRAY",
      "count":"P4",
      "dz":"P3*0.015",
      "da_deg":"360/P4",
      "note":"Slightly staggered vanes for aero benefit" },
    { "action":"SUBTRACT", "shape":"cylinder",
      "radius":"P1/2 - V1*1.5", "height":"P3+10", "z":"-5",
      "note":"Inner bore — removes vane protrusions on inner surface" }
  ]
}
"""

# ══════════════════════════════════════════════════════════════════════════════
#  SYSTEM PROMPT
# ══════════════════════════════════════════════════════════════════════════════

_SYSTEM_PROMPT = f"""You are the Siraal AI CAD Architect v3.0.
Your ONLY job: convert a natural language part description into a Siraal JSON CSG recipe
that the AutoCAD COM engine will execute. You have access to an enterprise-grade schema.

{_SCHEMA}

STUDIED EXAMPLES — learn exact structure, field names, and ordering:
{_EXAMPLES}

OUTPUT RULES (non-negotiable):
1. Output ONLY the raw JSON object. No markdown. No backticks. No explanation.
2. JSON must start with {{ and end with }}.
3. Every geometry step needs "action" and "shape". Modifier steps need only "action".
4. Step 0 MUST be action=BASE.
5. SET_VAR steps go BEFORE any step that uses the variable.
6. SUBTRACT tools: z="-5", height="dim+10". No exceptions.
7. Pattern rule: one template step, then pattern action immediately after.
8. INTERSECT, SLICE, SCALE, SWEEP come before PATTERN/MIRROR.
9. PATTERN/MIRROR/FILLET/CHAMFER come last.
10. Choose shapes intelligently:
    - axisymmetric body → revolve
    - hollow pipe/tube → pipe
    - hex nut/bolt head → polygon_prism (sides=6)
    - coil/spring → spring
    - odd volume = sphere ∩ box → BASE sphere + INTERSECT box
    - tapered body → frustum
    - swept channel → SWEEP
    - grid of features → ARRAY_GRID
    - non-uniform angles → PATTERN_RADIAL
    - rising helix of features → HELIX_ARRAY
"""

# ══════════════════════════════════════════════════════════════════════════════
#  VALIDATION ENGINE  v3.0
# ══════════════════════════════════════════════════════════════════════════════

_SHAPE_FIELDS: dict = {
    "cylinder":        {"radius", "height"},
    "box":             {"length", "width", "height"},
    "sphere":          {"radius"},
    "cone":            {"radius", "height"},
    "torus":           {"major_radius", "minor_radius"},
    "frustum":         {"radius_bottom", "radius_top", "height"},
    "pipe":            {"outer_radius", "inner_radius", "height"},
    "polygon_prism":   {"sides", "radius", "height"},
    "ellipsoid":       {"rx", "ry", "rz"},
    "spring":          {"coil_radius", "wire_radius", "pitch", "turns"},
    "extrude_profile": {"points", "height"},
    "revolve":         {"profile_points"},
}

_GEOMETRY_ACTIONS = {"BASE", "ADD", "SUBTRACT", "INTERSECT"}
_MODIFIER_ACTIONS = {
    "FILLET", "CHAMFER",
    "PATTERN_CIRCULAR", "PATTERN_LINEAR", "PATTERN_RADIAL",
    "ARRAY_GRID", "HELIX_ARRAY",
    "MIRROR", "SCALE", "SLICE", "SWEEP", "SET_VAR",
}
_ALL_ACTIONS = _GEOMETRY_ACTIONS | _MODIFIER_ACTIONS

_EVAL_SAFE_NS = {
    "__builtins__": None,
    "math": math, "pi": math.pi,
    "abs": abs, "min": min, "max": max, "round": round, "pow": pow,
    "sqrt": math.sqrt, "sin": math.sin, "cos": math.cos, "tan": math.tan,
    "asin": math.asin, "acos": math.acos, "atan": math.atan, "atan2": math.atan2,
    "hypot": math.hypot, "ceil": math.ceil, "floor": math.floor,
    "log": math.log, "log10": math.log10, "exp": math.exp,
    "radians": math.radians, "degrees": math.degrees,
}


def _can_eval(expr: str) -> bool:
    """Return True if expr evaluates to a number with dummy P1-P4=1, V1-V9=1."""
    ns = dict(_EVAL_SAFE_NS)
    ns.update({f"P{i}": 1.0 for i in range(1, 5)})
    ns.update({f"V{i}": 1.0 for i in range(1, 10)})
    try:
        float(eval(str(expr), ns))
        return True
    except Exception:
        return False


def validate_recipe(recipe: dict) -> list[str]:
    """
    Validates a recipe dict.  Returns a list of error strings.
    Empty list = valid.  Covers all v3.0 actions and shapes.
    """
    errors: list[str] = []
    if not isinstance(recipe, dict):
        return ["Recipe is not a JSON object"]
    if "Steps" not in recipe:
        return ["Missing 'Steps' array"]
    steps = recipe["Steps"]
    if not isinstance(steps, list) or len(steps) == 0:
        return ["'Steps' must be a non-empty array"]

    # Step 0 must be BASE (SET_VAR steps allowed before BASE)
    first_non_setvar = next((s for s in steps
                              if str(s.get("action","")).upper() != "SET_VAR"), None)
    if first_non_setvar and str(first_non_setvar.get("action","")).upper() != "BASE":
        errors.append(f"First geometry step must be BASE, got "
                       f"'{first_non_setvar.get('action')}'")

    has_base      = False
    defined_vars  = {"P1","P2","P3","P4"}

    for i, step in enumerate(steps):
        action = str(step.get("action", "")).upper()
        ctx    = f"Step {i} ({action})"

        if action not in _ALL_ACTIONS:
            errors.append(f"{ctx}: unknown action '{step.get('action')}'. "
                           f"Valid: {sorted(_ALL_ACTIONS)}")
            continue

        if action == "BASE":
            has_base = True

        # ── SET_VAR ──────────────────────────────────────────────────────────
        if action == "SET_VAR":
            var = str(step.get("var", ""))
            if not var:
                errors.append(f"{ctx}: missing 'var' field")
            elif not var.startswith(("V","P")):
                errors.append(f"{ctx}: var='{var}' should be V1-V9 or Pn")
            expr = step.get("expr", "")
            if expr and not _can_eval(str(expr)):
                errors.append(f"{ctx}: expr='{expr}' is not a valid expression")
            if var: defined_vars.add(var)
            continue

        # ── Geometry steps ────────────────────────────────────────────────────
        if action in _GEOMETRY_ACTIONS:
            if action == "SWEEP":
                if "profile_points" not in step:
                    errors.append(f"{ctx}: SWEEP needs 'profile_points'")
                if "path_points" not in step:
                    errors.append(f"{ctx}: SWEEP needs 'path_points'")
                continue

            shape = str(step.get("shape", "")).lower()
            if shape not in _SHAPE_FIELDS:
                errors.append(f"{ctx}: unknown shape '{shape}'. "
                               f"Valid: {sorted(_SHAPE_FIELDS)}")
                continue

            required = _SHAPE_FIELDS[shape]
            for field in required:
                if field not in step:
                    errors.append(f"{ctx} shape={shape}: missing '{field}'")
                    continue
                val = step[field]
                if isinstance(val, (int, float, list)):
                    continue
                if not _can_eval(str(val)):
                    errors.append(f"{ctx} '{field}'='{val}': not a valid expression")

            if action == "SUBTRACT":
                z_val = step.get("z", 0)
                if str(z_val) in ("0", "0.0") or z_val == 0:
                    errors.append(f"{ctx}: z=0 may miss — use z='-5'")

        # ── Modifier-specific checks ──────────────────────────────────────────
        elif action == "PATTERN_CIRCULAR":
            if "count" not in step:
                errors.append(f"{ctx}: needs 'count'")
        elif action == "PATTERN_LINEAR":
            for f in ("count", "dx", "dy", "dz"):
                if f not in step:
                    errors.append(f"{ctx}: needs '{f}'")
        elif action == "PATTERN_RADIAL":
            if "angles" not in step:
                errors.append(f"{ctx}: needs 'angles' list of degrees")
            elif not isinstance(step["angles"], list):
                errors.append(f"{ctx}: 'angles' must be a JSON list")
        elif action == "ARRAY_GRID":
            for f in ("rows", "cols", "dx", "dy"):
                if f not in step:
                    errors.append(f"{ctx}: needs '{f}'")
        elif action == "HELIX_ARRAY":
            if "count" not in step:
                errors.append(f"{ctx}: needs 'count'")
            if "dz" not in step:
                errors.append(f"{ctx}: needs 'dz'")
        elif action == "MIRROR":
            plane = str(step.get("plane", "")).upper()
            if plane not in ("XZ", "YZ", "XY"):
                errors.append(f"{ctx}: plane='{plane}' must be XZ, YZ, or XY")
        elif action == "SCALE":
            if "factor" not in step:
                errors.append(f"{ctx}: needs 'factor'")
        elif action == "SLICE":
            if "plane_normal" not in step:
                errors.append(f"{ctx}: needs 'plane_normal' (X|Y|Z)")
        elif action == "FILLET":
            if "radius" not in step:
                errors.append(f"{ctx}: needs 'radius'")
        elif action == "CHAMFER":
            if "distance" not in step:
                errors.append(f"{ctx}: needs 'distance'")

    if not has_base:
        errors.append("No BASE step found")

    return errors


# ══════════════════════════════════════════════════════════════════════════════
#  GEMINI CALLER  (with one validation-feedback retry)
# ══════════════════════════════════════════════════════════════════════════════

def _call_gemini(prompt: str, api_key: Optional[str],
                 model: str, log_cb: Callable) -> Optional[dict]:
    """
    Call Gemini, parse the JSON, validate it.
    On validation failure, make one more call with the errors fed back.
    Returns a valid recipe dict or None.
    """
    if not _GENAI_OK:
        log_cb("[AI] ✘ google-genai not installed. Run: pip install google-genai")
        return None

    client = genai.Client(api_key=api_key) if api_key else genai.Client()

    def _one_call(contents: str) -> Optional[dict]:
        try:
            resp = client.models.generate_content(
                model=model,
                contents=contents,
                config=gtypes.GenerateContentConfig(
                    system_instruction=_SYSTEM_PROMPT,
                    temperature=0.1,
                    max_output_tokens=4096,
                )
            )
            raw = resp.text.strip()
            # Strip accidental markdown fences
            raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
            raw = re.sub(r"\s*```$", "", raw)
            return json.loads(raw)
        except json.JSONDecodeError as e:
            log_cb(f"[AI] ✘ JSON parse error: {e}")
            log_cb(f"[AI]   Raw (first 400 chars): {resp.text[:400]}")
            return None
        except Exception as e:
            log_cb(f"[AI] ✘ API error: {e}")
            return None

    # ── Attempt 1 ────────────────────────────────────────────────────────────
    log_cb(f"[AI] Calling Gemini ({model})…")
    recipe = _one_call(prompt)
    if recipe is None:
        return None

    errors = validate_recipe(recipe)
    if not errors:
        log_cb(f"[AI] ✔ Recipe valid ({len(recipe.get('Steps', []))} steps)")
        return recipe

    # ── Attempt 2: feed errors back ──────────────────────────────────────────
    log_cb(f"[AI] ⚠ Validation found {len(errors)} issue(s) — retrying with corrections…")
    for e in errors:
        log_cb(f"     • {e}")

    correction_prompt = (
        f"{prompt}\n\n"
        f"Your previous attempt had these validation errors:\n"
        + "\n".join(f"  - {e}" for e in errors)
        + "\n\nFix ALL of these issues and return the corrected JSON only."
    )
    recipe2 = _one_call(correction_prompt)
    if recipe2 is None:
        return None

    errors2 = validate_recipe(recipe2)
    if errors2:
        log_cb(f"[AI] ⚠ Still {len(errors2)} issue(s) after retry — saving anyway:")
        for e in errors2:
            log_cb(f"     • {e}")
    else:
        log_cb(f"[AI] ✔ Corrected recipe valid ({len(recipe2.get('Steps', []))} steps)")
    return recipe2


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN PUBLIC FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def generate_siraal_shape(
        part_name:   str,
        description: str,
        api_key:     Optional[str] = None,
        model_name:  str = "gemini-2.5-flash",
        log_cb:      Callable = print,
) -> bool:
    """
    Generate a Siraal JSON template from a natural language description
    and save it to templates/{part_name}.json.

    Parameters
    ──────────
    part_name   — used as the filename and Part_Type in the BOM
                  (do NOT prefix with Custom_ — the engine handles lookup)
    description — plain English description of the part, mentioning
                  what P1, P2, P3, P4 mean for this specific part
    api_key     — Gemini API key (or set GEMINI_API_KEY env var)
    model_name  — Gemini model to use
    log_cb      — logging callback (use print or GUI log function)

    Returns True on success, False on failure.
    """
    log_cb(f"\n[AI] ═══════════════════════════════════════════════")
    log_cb(f"[AI]  Generating template for: {part_name}")
    log_cb(f"[AI] ═══════════════════════════════════════════════")

    # Build the user prompt
    user_prompt = (
        f"Build a Siraal JSON CSG recipe for the following part.\n\n"
        f"Part Name : {part_name}\n"
        f"Description: {description}\n\n"
        f"Use the schema and examples above. "
        f"Output only the raw JSON object."
    )

    recipe = _call_gemini(user_prompt, api_key, model_name, log_cb)
    if recipe is None:
        log_cb(f"[AI] ✘ Failed to generate recipe for '{part_name}'")
        return False

    # Save
    os.makedirs("templates", exist_ok=True)
    file_path = os.path.join("templates", f"{part_name}.json")
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(recipe, f, indent=2)

    log_cb(f"[AI] ✔ Saved to {file_path}")
    log_cb(f"[AI]   Steps: {len(recipe.get('Steps', []))}")
    log_cb(f"[AI]   Add Part_Type='{part_name}' to your BOM to build it.")
    return True


def generate_and_queue(
        part_name:   str,
        description: str,
        material:    str  = "Steel-4140",
        p1: float = 100.0,
        p2: float = 30.0,
        p3: float = 50.0,
        p4: float = 20.0,
        qty:         int  = 1,
        priority:    str  = "High",
        excel_path:  Optional[str] = None,
        api_key:     Optional[str] = None,
        model_name:  str  = "gemini-2.5-flash",
        log_cb:      Callable = print,
) -> bool:
    """
    Full pipeline:
      1. Generate the JSON template
      2. Validate it
      3. Append a new row to the BOM Excel (if excel_path is given)

    After this returns True, just run the engine on the Excel and the
    custom part will be built automatically via the template lookup.

    Parameters (geometry)
    ──────────────────────
    p1-p4  — the four parameter values for this part (meaning depends on
               your description, e.g. P1=OD, P2=bore, P3=height, P4=slots)
    """
    # Step 1 — generate template
    ok = generate_siraal_shape(part_name, description, api_key, model_name, log_cb)
    if not ok:
        return False

    # Step 2 — add to BOM if requested
    if excel_path and os.path.exists(excel_path):
        ok2 = _append_to_bom(
            excel_path, part_name, material, p1, p2, p3, p4, qty, priority, log_cb
        )
        if ok2:
            log_cb(f"[AI] ✔ Part '{part_name}' queued in BOM — run engine to build.")
        else:
            log_cb(f"[AI] ⚠ Template saved but BOM update failed. Add manually.")
    else:
        log_cb(f"[AI] ℹ No excel_path given — add Part_Type='{part_name}' to BOM manually.")

    return True


def _append_to_bom(
        excel_path: str,
        part_name:  str,
        material:   str,
        p1: float, p2: float, p3: float, p4: float,
        qty: int, priority: str,
        log_cb: Callable,
) -> bool:
    """
    Appends one row to the BOM_Gears sheet using openpyxl.
    Preserves all existing rows, formulas, and styling.
    Regenerates the mass/cost Excel formulas for the new row.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        log_cb("[AI] ✘ openpyxl not installed. Run: pip install openpyxl")
        return False

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)

        # Find the sheet
        sheet_name = None
        for candidate in ["BOM_Gears", "BOM_3D", "BOM", "Parts"]:
            if candidate in wb.sheetnames:
                sheet_name = candidate; break
        if sheet_name is None:
            log_cb(f"[AI] ✘ No BOM sheet found in {excel_path}")
            return False
        ws = wb[sheet_name]

        # Find the last data row (before the TOTALS sentinel)
        DATA_START = 4
        last_data_row = DATA_START - 1
        for r in range(DATA_START, ws.max_row + 1):
            val = ws.cell(row=r, column=2).value
            if val is None or str(val).lower().startswith("total"):
                break
            last_data_row = r

        new_row  = last_data_row + 1
        row_idx  = new_row - DATA_START + 1

        # Build the auto-generated part number
        part_no  = f"AI-{part_name.upper()[:12]}-{row_idx:03d}"

        # Density lookup for formula (approximate — engine recalculates anyway)
        _densities = {
            "Steel-1020": 7.87e-3, "Steel-4140": 7.85e-3, "Steel-EN36": 7.85e-3,
            "Al-6061": 2.70e-3,    "Al-7075": 2.81e-3,    "Brass-C360": 8.50e-3,
            "Nylon-66": 1.14e-3,   "Ti-6Al-4V": 4.43e-3,  "Cast-Iron": 7.20e-3,
        }
        _costs = {
            "Steel-1020": 125, "Steel-4140": 185, "Steel-EN36": 220,
            "Al-6061": 265,    "Al-7075": 390,    "Brass-C360": 520,
            "Nylon-66": 415,   "Ti-6Al-4V": 3800, "Cast-Iron": 95,
        }
        density  = _densities.get(material, 7.85e-3)
        cost_pkg = _costs.get(material, 185)

        E = f"E{new_row}"; F = f"F{new_row}"
        G = f"G{new_row}"; H = f"H{new_row}"

        cells = [
            (1,  row_idx),                    # A — index
            (2,  part_no),                    # B — part number
            (3,  part_name),                  # C — Part_Type (triggers template lookup)
            (4,  material),                   # D
            (5,  int(p1) if p1 == int(p1) else p1),  # E — Param_1
            (6,  p2),                         # F — Param_2
            (7,  p3),                         # G — Param_3
            (8,  p4),                         # H — Param_4
            (9,  qty),                        # I — Qty
            (10, priority),                   # J — Priority
            (11, "YES"),                      # K — Enabled
            (12, f"AI-generated: {part_name}"),# L — Description
            # M — mass formula
            (13, f"=ROUND(PI()*(({E}*{F}/2+{F})^2-({H}/2)^2)*{G}*{density:.8f},3)"),
            # N — cost formula
            (14, f"=ROUND(M{new_row}*{cost_pkg},2)"),
            (15, f"Template: templates/{part_name}.json"),  # O — Notes
        ]
        for col, value in cells:
            ws.cell(row=new_row, column=col, value=value)

        # Rewrite TOTALS row
        totals_row = new_row + 1
        ws.cell(row=totals_row, column=2, value="TOTALS (all enabled):")
        for col_l in ["M", "N"]:
            ws.cell(row=totals_row, column={"M":13,"N":14}[col_l],
                    value=f"=SUMIF(K{DATA_START}:K{new_row},\"YES\","
                           f"{col_l}{DATA_START}:{col_l}{new_row})")

        wb.save(excel_path)
        log_cb(f"[AI] ✔ Row {new_row} added to '{sheet_name}': {part_no}")
        return True

    except Exception as e:
        log_cb(f"[AI] ✘ BOM update failed: {e}")
        import traceback; log_cb(traceback.format_exc())
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  STANDALONE TEST
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys

    print("=" * 60)
    print("  SIRAAL AI SHAPE COMPILER  v2.0")
    print("=" * 60)

    api_key = os.environ.get("GEMINI_API_KEY", "")
    if len(sys.argv) > 1:
        api_key = sys.argv[1]
    if not api_key:
        print("Pass your Gemini key as:  python genai_creator.py YOUR_KEY")
        sys.exit(1)

    # ── EXAMPLE TEST PROMPT — paste into your GUI or run directly ────────────
    generate_and_queue(
        part_name   = "Turbine_Disc_LT",
        description = (
            "A forged high-pressure turbine disc for a gas turbine engine. "
            "P1 = outer diameter in mm (e.g. 300). "
            "P2 = central bore diameter in mm (e.g. 60). "
            "P3 = disc thickness in mm (e.g. 45). "
            "P4 = number of fir-tree blade slots around the rim (e.g. 36). "
            "The disc has: "
            "1. A large central bore for the shaft. "
            "2. Fir-tree shaped slots evenly spaced around the outer rim "
            "   (rectangular approximation: width = OD*0.07, depth = OD*0.06) "
            "   to hold turbine blades. "
            "3. Eight equally-spaced balance holes at 70% of the disc radius. "
            "4. A 3mm fillet at all internal transitions for stress relief. "
            "Material will be Steel-4140 (will be set via BOM)."
        ),
        material    = "Steel-4140",
        p1 = 300,   # OD
        p2 = 60,    # Bore
        p3 = 45,    
        p4 = 36,    
        qty         = 1,
        priority    = "High",
        excel_path  = "excels/demo_gears_3d.xlsx",
        api_key     = api_key,
    )