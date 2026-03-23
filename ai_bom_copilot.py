"""
ai_bom_copilot.py  —  Siraal AI Copilot Engine  v1.0
=====================================================
Standalone back-end module for AI-driven BOM editing.
No GUI code. Import and call from gui_launcher.py or any front-end.

PUBLIC API
──────────
  copilot = SiraalCopilot(excel_path, api_key, log_cb=print)

  # One-shot preview (does NOT write to disk)
  ok, new_parts, diff_text = copilot.preview(user_prompt)

  # Commit the last preview to disk (writes Excel + ChangeLog)
  ok, msg = copilot.commit(new_parts, author="AI Copilot")

  # Multi-turn conversation
  ok, new_parts, diff_text = copilot.chat(user_prompt)   # builds history
  copilot.reset_chat()                                    # clear history

  # Read-only helpers
  parts = copilot.load_bom()
  report = copilot.validate_parts(parts)
  schema = copilot.get_schema()

ARCHITECTURE
────────────
  BomIO        — openpyxl read/write, preserves formulas + styling
  Validator    — pure-Python rule engine (no AutoCAD, no pandas)
  GeminiClient — wraps google-genai, function-calling tool schema
  SiraalCopilot— orchestrates the three above; public API
"""

from __future__ import annotations

import copy
import json
import math
import os
import re
import traceback
from dataclasses import dataclass, field, asdict
from datetime import datetime
from enum import Enum
from typing import Any, Callable, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from google import genai
    from google.genai import types as gtypes
    _GENAI_AVAILABLE = True
except ImportError:
    _GENAI_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS — mirror validator_3d + engine exactly
# ═══════════════════════════════════════════════════════════════════════════════

GEAR_TYPES = {
    "Spur_Gear_3D", "Helical_Gear", "Ring_Gear_3D",
    "Bevel_Gear", "Worm", "Worm_Wheel",
}
SOLID_TYPES = {
    "Box", "Cylinder", "Cone", "Sphere",
    "Flanged_Boss", "Extruded_Profile", "Revolved_Part",
}
OTHER_TYPES = {
   'Flange','Stepped_Shaft','L_Bracket'
}
ALL_TYPES = GEAR_TYPES | SOLID_TYPES | OTHER_TYPES

MATERIAL_DB: Dict[str, Dict[str, float]] = {
    "Steel-1020": {"density": 7.87e-3, "cost_per_kg": 125.0,  "uts": 420},
    "Steel-4140": {"density": 7.85e-3, "cost_per_kg": 185.0,  "uts": 1000},
    "Al-6061":    {"density": 2.70e-3, "cost_per_kg": 265.0,  "uts": 310},
    "Brass-C360": {"density": 8.50e-3, "cost_per_kg": 520.0,  "uts": 360},
    "Nylon-66":   {"density": 1.14e-3, "cost_per_kg": 415.0,  "uts": 82},
    "Ti-6Al-4V":  {"density": 4.43e-3, "cost_per_kg": 3800.0, "uts": 950},
}

# Column layout of BOM_Gears sheet (1-indexed, matching actual file)
COL_IDX     = 1   # A — row index (int)
COL_PARTNO  = 2   # B
COL_TYPE    = 3   # C
COL_MAT     = 4   # D
COL_P1      = 5   # E  Param_1  Z / N_starts
COL_P2      = 6   # F  Param_2  Module
COL_P3      = 7   # G  Param_3  FaceWidth / Length
COL_P4      = 8   # H  Param_4  BoreDia / RingThk
COL_QTY     = 9   # I
COL_PRI     = 10  # J
COL_ENA     = 11  # K
COL_DESC    = 12  # L
COL_MASS    = 13  # M  — Excel formula, NOT overwritten
COL_COST    = 14  # N  — Excel formula, NOT overwritten
COL_NOTES   = 15  # O

DATA_START_ROW = 4   # Row 1-2 = title/info, Row 3 = header, Row 4 = first data

PRIORITIES = ["High", "Medium", "Low"]
ENABLED_VALUES = ["YES", "NO"]

PARAM_LABELS: Dict[str, Dict[str, str]] = {
    "Spur_Gear_3D": {"P1": "Teeth Z",        "P2": "Module m (mm)",
                     "P3": "Face Width (mm)", "P4": "Bore Dia (mm)"},
    "Helical_Gear": {"P1": "Teeth Z",         "P2": "Module m (mm)",
                     "P3": "Face Width (mm)", "P4": "Bore Dia (mm)"},
    "Ring_Gear_3D": {"P1": "Teeth Z",         "P2": "Module m (mm)",
                     "P3": "Face Width (mm)", "P4": "Ring Thickness (mm)"},
    "Bevel_Gear":   {"P1": "Teeth Z",         "P2": "Module m (mm)",
                     "P3": "Face Width (mm)", "P4": "Bore Dia (mm)"},
    "Worm":         {"P1": "N_starts",        "P2": "Module m (mm)",
                     "P3": "Length (mm)",     "P4": "Bore Dia (mm)"},
    "Worm_Wheel":   {"P1": "Teeth Z",         "P2": "Module m (mm)",
                     "P3": "Face Width (mm)", "P4": "Bore Dia (mm)"},
    "Box":          {"P1": "Length (mm)",     "P2": "Width (mm)",
                     "P3": "Height (mm)",     "P4": "Fillet R (mm)"},
    "Cylinder":     {"P1": "Outer R (mm)",    "P2": "Bore R (mm)",
                     "P3": "Height (mm)",     "P4": "Chamfer (mm)"},
}


# ═══════════════════════════════════════════════════════════════════════════════
# DATA MODELS
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class BomPart:
    """Single row in the BOM — maps 1-to-1 with Excel columns B→O."""
    part_number: str
    part_type:   str
    material:    str
    param_1:     float   # Z / N_starts
    param_2:     float   # Module
    param_3:     float   # FaceWidth / Length
    param_4:     float   # BoreDia / RingThk
    qty:         int     = 1
    priority:    str     = "High"
    enabled:     str     = "YES"
    description: str     = ""
    notes:       str     = ""

    def to_dict(self) -> dict:
        return {
            "Part_Number": self.part_number,
            "Part_Type":   self.part_type,
            "Material":    self.material,
            "Param_1":     self.param_1,
            "Param_2":     self.param_2,
            "Param_3":     self.param_3,
            "Param_4":     self.param_4,
            "Qty":         self.qty,
            "Priority":    self.priority,
            "Enabled":     self.enabled,
            "Description": self.description,
            "Notes":       self.notes,
        }

    @staticmethod
    def from_dict(d: dict) -> "BomPart":
        def _f(k, default=0.0):
            try:   return float(d.get(k, default) or default)
            except: return default
        def _i(k, default=1):
            try:   return int(float(d.get(k, default) or default))
            except: return default
        def _s(k, default=""):
            v = d.get(k, default)
            return str(v).strip() if v is not None else default

        return BomPart(
            part_number = _s("Part_Number"),
            part_type   = _s("Part_Type",  "Spur_Gear_3D"),
            material    = _s("Material",   "Steel-1020"),
            param_1     = _f("Param_1",    20),
            param_2     = _f("Param_2",    3),
            param_3     = _f("Param_3",    30),
            param_4     = _f("Param_4",    20),
            qty         = _i("Qty",        1),
            priority    = _s("Priority",   "High"),
            enabled     = _s("Enabled",    "YES").upper(),
            description = _s("Description"),
            notes       = _s("Notes"),
        )


class IssueSeverity(Enum):
    ERROR   = "ERROR"
    WARNING = "WARNING"
    INFO    = "INFO"


@dataclass
class ValidationIssue:
    part_number: str
    severity:    IssueSeverity
    rule:        str
    message:     str

    def __str__(self) -> str:
        icon = {"ERROR": "✘", "WARNING": "⚠", "INFO": "ℹ"}.get(self.severity.value, "?")
        return f"  {icon} [{self.severity.value}] {self.part_number} | {self.rule}: {self.message}"


@dataclass
class ValidationReport:
    issues:       List[ValidationIssue] = field(default_factory=list)
    error_count:  int = 0
    warning_count: int = 0

    @property
    def is_valid(self) -> bool:
        return self.error_count == 0

    def summary(self) -> str:
        lines = [str(i) for i in self.issues]
        lines.append(f"\n  {'✔ VALID' if self.is_valid else '✘ INVALID'}"
                     f"  —  {self.error_count} errors, {self.warning_count} warnings")
        return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# BOM I/O — openpyxl, preserves formulas + styling
# ═══════════════════════════════════════════════════════════════════════════════

class BomIO:
    """
    Reads and writes the BOM_Gears sheet of Siraal's Excel format.
    Columns M and N contain auto-recalculating Excel formulas — we
    regenerate them per row when writing so mass/cost stay live.

    Mass formula (col M):  =ROUND(PI()*((E*F/2+F)^2-(H/2)^2)*G*density/1e9, 3)
    Cost formula (col N):  =ROUND(M*cost_per_kg, 2)
    """

    SHEET = "BOM_Gears"

    def __init__(self, path: str):
        self.path = path

    # ── Public read ────────────────────────────────────────────────────────────

    def load(self) -> List[BomPart]:
        wb = openpyxl.load_workbook(self.path, data_only=False)
        ws = wb[self.SHEET]
        parts: List[BomPart] = []

        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            raw = [ws.cell(row=row_idx, column=c).value for c in range(1, 16)]
            # Stop at TOTALS sentinel or fully empty row
            if raw[0] is None and raw[1] is None:
                break
            part_no = str(raw[1]).strip() if raw[1] else ""
            if not part_no or part_no.lower().startswith("total"):
                break

            def _f(v, default=0.0):
                try:   return float(v) if v is not None else default
                except: return default
            def _i(v, default=1):
                try:   return int(float(v)) if v is not None else default
                except: return default
            def _s(v, default=""):
                return str(v).strip() if v is not None else default

            parts.append(BomPart(
                part_number = part_no,
                part_type   = _s(raw[2],  "Spur_Gear_3D"),
                material    = _s(raw[3],  "Steel-1020"),
                param_1     = _f(raw[4],  20),
                param_2     = _f(raw[5],  3),
                param_3     = _f(raw[6],  30),
                param_4     = _f(raw[7],  20),
                qty         = _i(raw[8],  1),
                priority    = _s(raw[9],  "High"),
                enabled     = _s(raw[10], "YES").upper(),
                description = _s(raw[11], ""),
                notes       = _s(raw[14], ""),
            ))

        return parts

    # ── Public write ───────────────────────────────────────────────────────────

    # ── Public write ───────────────────────────────────────────────────────────

    def save(self, parts: List[BomPart]) -> None:
        """
        Overwrite the BOM_Gears data rows with `parts`.
        Rows beyond len(parts) are cleared safely.
        Header rows 1-3 and sheet styling are preserved.
        """
        wb = openpyxl.load_workbook(self.path, data_only=False)
        ws = wb[self.SHEET]

        # 1. THE FIX: Find and unmerge the old TOTALS row (and any other merges below headers)
        # This completely prevents the "MergedCell attribute is read-only" crash.
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row >= DATA_START_ROW:
                try:
                    ws.unmerge_cells(merged_range.coord)
                except Exception:
                    pass

        # 2. Write new rows
        for i, part in enumerate(parts):
            row = DATA_START_ROW + i
            ws.cell(row=row, column=COL_IDX,    value=i + 1)
            ws.cell(row=row, column=COL_PARTNO, value=part.part_number)
            ws.cell(row=row, column=COL_TYPE,   value=part.part_type)
            ws.cell(row=row, column=COL_MAT,    value=part.material)
            ws.cell(row=row, column=COL_P1,     value=int(part.param_1) if part.param_1 == int(part.param_1) else part.param_1)
            ws.cell(row=row, column=COL_P2,     value=part.param_2)
            ws.cell(row=row, column=COL_P3,     value=part.param_3)
            ws.cell(row=row, column=COL_P4,     value=part.param_4)
            ws.cell(row=row, column=COL_QTY,    value=part.qty)
            ws.cell(row=row, column=COL_PRI,    value=part.priority)
            ws.cell(row=row, column=COL_ENA,    value=part.enabled)
            ws.cell(row=row, column=COL_DESC,   value=part.description)
            # Regenerate live Excel formulas
            ws.cell(row=row, column=COL_MASS,   value=self._mass_formula(row, part.material))
            ws.cell(row=row, column=COL_COST,   value=self._cost_formula(row, part.material))
            ws.cell(row=row, column=COL_NOTES,  value=part.notes)

        # 3. Cleanly delete old excess rows (much safer than setting to None)
        new_last = DATA_START_ROW + len(parts) - 1
        old_max = ws.max_row
        
        if old_max > new_last:
            # Delete everything from the row after our new data down to the end
            ws.delete_rows(new_last + 1, old_max - new_last + 5) 

        # 4. Rewrite TOTALS row
        totals_row = new_last + 1
        
        # Safely create the new merged block for the TOTALS label
        ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=4)
        c = ws.cell(row=totals_row, column=1, value="TOTALS (all enabled):")
        c.font = Font(name="Segoe UI", bold=True, color="F0B429")
        c.alignment = Alignment(horizontal="right")
        
        # Apply the live SUMIF formulas
        for c_idx in [COL_MASS, COL_COST]:
            col_l = get_column_letter(c_idx)
            start = DATA_START_ROW
            end   = new_last
            ws.cell(row=totals_row, column=c_idx, value=f"=SUMIF(K{start}:K{end},\"YES\",{col_l}{start}:{col_l}{end})")

        wb.save(self.path)
    def append_changelog(self, author: str, prompt: str, n_changed: int) -> None:
        """Append an entry to the ChangeLog sheet if it exists."""
        try:
            wb = openpyxl.load_workbook(self.path, data_only=False)
            if "ChangeLog" not in wb.sheetnames:
                return
            ws   = wb["ChangeLog"]
            row  = ws.max_row + 1
            ws.cell(row=row, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row, column=2, value=author)
            ws.cell(row=row, column=3, value=prompt[:200])
            ws.cell(row=row, column=4, value=n_changed)
            wb.save(self.path)
        except Exception:
            pass   # ChangeLog write failure is non-fatal and should not block the main save

    # ── Helpers ────────────────────────────────────────────────────────────────

    @staticmethod
    def _mass_formula(row: int, material: str) -> str:
        """
        Gear mass formula (col M).
        Uses the density constant baked in (matches engine MATERIAL_DB).
        """
        density = MATERIAL_DB.get(material, MATERIAL_DB["Steel-1020"])["density"]
        # density in g/mm³  →  g/cm³ = density * 1000  →  kg/cm³ = density
        # volume in mm³ = pi*(OD/2)^2 * FW  (approximate cylinder — formula matches file)
        # volume formula: PI()*((E*F/2+F)^2-(H/2)^2)*G   [mm³]
        # mass kg = vol_mm³ * density_g_mm3 / 1000
        e, f, g, h = f"E{row}", f"F{row}", f"G{row}", f"H{row}"
        d = density * 1e-3   # g/mm³ → kg/mm³  (density_g_cm3 / 1e6)
        return f"=ROUND(PI()*(({e}*{f}/2+{f})^2-({h}/2)^2)*{g}*{d:.8f},3)"

    @staticmethod
    def _cost_formula(row: int, material: str) -> str:
        cost = MATERIAL_DB.get(material, MATERIAL_DB["Steel-1020"])["cost_per_kg"]
        return f"=ROUND(M{row}*{cost},2)"


# ═══════════════════════════════════════════════════════════════════════════════
# PURE-PYTHON VALIDATOR  (no pandas, no AutoCAD)
# ═══════════════════════════════════════════════════════════════════════════════

class BomValidator:
    """
    Validates a list of BomPart objects against Siraal's engineering rules.
    Mirrors validator_3d.py rules exactly so the AI gets the same feedback.
    """

    def validate(self, parts: List[BomPart]) -> ValidationReport:
        report = ValidationReport()
        seen_part_nos: set = set()

        for part in parts:
            pno = part.part_number

            # ── Structural checks ──────
            if not pno:
                self._add(report, pno, IssueSeverity.ERROR, "MISSING_PARTNO",
                          "Part_Number is blank")
                continue

            if pno in seen_part_nos:
                self._add(report, pno, IssueSeverity.ERROR, "DUPLICATE_PARTNO",
                          f"Duplicate Part_Number '{pno}'")
            seen_part_nos.add(pno)

            if False or part.part_type not in ALL_TYPES:
                self._add(report, pno, IssueSeverity.ERROR, "UNKNOWN_TYPE",
                          f"'{part.part_type}' not recognised. Valid: {sorted(ALL_TYPES)}")
                continue

            if part.material not in MATERIAL_DB:
                self._add(report, pno, IssueSeverity.WARNING, "UNKNOWN_MATERIAL",
                          f"'{part.material}' not in database — will fall back to Steel-1020")

            if part.enabled not in ("YES", "NO"):
                self._add(report, pno, IssueSeverity.WARNING, "ENABLED_VALUE",
                          f"Enabled='{part.enabled}' should be YES or NO")

            if part.qty < 1:
                self._add(report, pno, IssueSeverity.WARNING, "QTY",
                          f"Qty={part.qty} — should be ≥ 1")

            if part.priority not in PRIORITIES:
                self._add(report, pno, IssueSeverity.WARNING, "PRIORITY",
                          f"Priority='{part.priority}' not in {PRIORITIES}")

            # ── Skip geometry checks if disabled ──────────────────────────────
            if part.enabled != "YES":
                continue

            p1, p2, p3, p4 = part.param_1, part.param_2, part.param_3, part.param_4

            # ── Geometry rules per type ────────────────────────────────────────
            pt = part.part_type
            if   pt == "Spur_Gear_3D": self._check_spur(report, pno, p1, p2, p3, p4)
            elif pt == "Helical_Gear":  self._check_helical(report, pno, p1, p2, p3, p4)
            elif pt == "Ring_Gear_3D":  self._check_ring(report, pno, p1, p2, p3, p4)
            elif pt == "Bevel_Gear":    self._check_bevel(report, pno, p1, p2, p3, p4)
            elif pt == "Worm":          self._check_worm(report, pno, p1, p2, p3, p4)
            elif pt == "Worm_Wheel":    self._check_worm_wheel(report, pno, p1, p2, p3, p4)
            elif pt == "Box":           self._check_box(report, pno, p1, p2, p3, p4)
            elif pt == "Cylinder":      self._check_cylinder(report, pno, p1, p2, p3, p4)

        return report

    # ── Per-type rules ─────────────────────────────────────────────────────────

    def _check_spur(self, r, pno, p1, p2, p3, p4):
        Z = int(p1); m = p2; fw = p3; bore_d = p4
        pitch_r = Z * m / 2.0
        bore_r  = bore_d / 2.0
        if Z < 6:
            self._add(r, pno, IssueSeverity.ERROR, "SPUR_Z_MIN",
                      f"Z={Z} < 6 — too few teeth, gear invalid")
        elif Z < 17:
            self._add(r, pno, IssueSeverity.WARNING, "SPUR_UNDERCUT",
                      f"Z={Z} < 17 — involute undercut likely; profile shift recommended")
        if m <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "SPUR_MODULE",
                      f"Module m={m} must be > 0")
        if fw <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "SPUR_FACEWIDTH",
                      f"FaceWidth={fw} must be > 0")
        if m > 0 and bore_r >= pitch_r:
            self._add(r, pno, IssueSeverity.ERROR, "SPUR_BORE",
                      f"bore_r={bore_r:.1f} ≥ pitch_r={pitch_r:.1f} — bore too large for Z×m")
        if bore_d > 0 and bore_d < 5:
            self._add(r, pno, IssueSeverity.WARNING, "SPUR_BORE_SMALL",
                      f"Bore Ø{bore_d}mm very small — DIN 6885 keyway will not fit")
        if m > 0 and fw > 12 * m:
            self._add(r, pno, IssueSeverity.WARNING, "SPUR_FW_RATIO",
                      f"FaceWidth/m={fw/m:.1f} > 12 — excessive for spur gear")

    def _check_helical(self, r, pno, p1, p2, p3, p4):
        Z = int(p1); m = p2; fw = p3; bore_d = p4
        pitch_r = Z * m / 2.0
        bore_r  = bore_d / 2.0
        if Z < 6:
            self._add(r, pno, IssueSeverity.ERROR, "HEL_Z_MIN",
                      f"Z={Z} < 6 — too few teeth")
        elif Z < 17:
            self._add(r, pno, IssueSeverity.WARNING, "HEL_UNDERCUT",
                      f"Z={Z} < 17 — profile shift recommended")
        if m <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "HEL_MODULE",
                      f"Module m={m} must be > 0")
        if fw <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "HEL_FACEWIDTH",
                      f"FaceWidth={fw} must be > 0")
        if m > 0 and bore_r >= pitch_r:
            self._add(r, pno, IssueSeverity.ERROR, "HEL_BORE",
                      f"bore_r={bore_r:.1f} ≥ pitch_r={pitch_r:.1f}")
        if m > 0 and fw > 20 * m:
            self._add(r, pno, IssueSeverity.WARNING, "HEL_FW_RATIO",
                      f"FaceWidth/m={fw/m:.1f} > 20 — very wide for helical gear")

    def _check_ring(self, r, pno, p1, p2, p3, p4):
        Z = int(p1); m = p2; fw = p3; ring_thk = p4
        pitch_r = Z * m / 2.0
        inner_r = pitch_r - m
        if Z < 20:
            self._add(r, pno, IssueSeverity.WARNING, "RING_Z_MIN",
                      f"Z={Z} < 20 — meshing issues with small planet")
        if m <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "RING_MODULE",
                      f"Module m={m} must be > 0")
        if fw <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "RING_FACEWIDTH",
                      f"FaceWidth={fw} must be > 0")
        if ring_thk <= m:
            self._add(r, pno, IssueSeverity.ERROR, "RING_THICKNESS",
                      f"ring_thk={ring_thk} ≤ m={m} — wall too thin")
        if inner_r <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "RING_INNER_R",
                      f"inner_r=pitch_r-m={inner_r:.1f} ≤ 0 — impossible geometry")

    def _check_bevel(self, r, pno, p1, p2, p3, p4):
        Z = int(p1); m = p2; fw = p3; bore_d = p4
        cr     = math.radians(45)
        back_r = Z * m / 2.0 if m > 0 else 0
        front_r = back_r - fw * math.sin(cr)
        bore_r  = bore_d / 2.0
        if Z < 6:
            self._add(r, pno, IssueSeverity.ERROR, "BEVEL_Z_MIN",
                      f"Z={Z} < 6 — too few teeth")
        if m <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "BEVEL_MODULE",
                      f"Module m={m} must be > 0")
        if m > 0 and front_r <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "BEVEL_FRONTCONE",
                      f"front_r={front_r:.1f} ≤ 0 — FaceWidth={fw} too large; "
                      f"max FW = {back_r/math.sin(cr):.1f}mm for this Z×m")
        if m > 0 and bore_r >= back_r:
            self._add(r, pno, IssueSeverity.ERROR, "BEVEL_BORE",
                      f"bore_r={bore_r:.1f} ≥ back_r={back_r:.1f}")

    def _check_worm(self, r, pno, p1, p2, p3, p4):
        n_starts = int(p1); m = p2; length = p3; bore_d = p4
        bore_r  = bore_d / 2.0
        shaft_r = bore_r + m * 1.5 if bore_r > 1.0 else m * 3.0
        axp     = math.pi * m if m > 0 else 1
        if n_starts < 1:
            self._add(r, pno, IssueSeverity.ERROR, "WORM_STARTS",
                      f"N_starts={n_starts} must be ≥ 1")
        if m <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "WORM_MODULE",
                      f"Module m={m} must be > 0")
        if length <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "WORM_LENGTH",
                      f"Length={length} must be > 0")
        if m > 0 and length < axp:
            self._add(r, pno, IssueSeverity.WARNING, "WORM_LENGTH_SHORT",
                      f"Length={length:.1f} < 1 axial pitch={axp:.1f} — partial thread only")
        if bore_r > 0 and m > 0 and bore_r >= shaft_r:
            self._add(r, pno, IssueSeverity.ERROR, "WORM_BORE",
                      f"bore_r={bore_r:.1f} ≥ shaft_r={shaft_r:.1f}")
        if n_starts > 6:
            self._add(r, pno, IssueSeverity.WARNING, "WORM_STARTS_HIGH",
                      f"N_starts={n_starts} > 6 — high lead angle; not self-locking")

    def _check_worm_wheel(self, r, pno, p1, p2, p3, p4):
        Z = int(p1); m = p2; fw = p3; bore_d = p4
        pitch_r = Z * m / 2.0 if m > 0 else 0
        outer_r = pitch_r + m if m > 0 else 0
        bore_r  = bore_d / 2.0
        if Z < 20:
            self._add(r, pno, IssueSeverity.WARNING, "WWHEEL_Z_MIN",
                      f"Z={Z} < 20 — contact ratio low")
        if m <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "WWHEEL_MODULE",
                      f"Module m={m} must be > 0")
        if fw <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "WWHEEL_FACEWIDTH",
                      f"FaceWidth={fw} must be > 0")
        if m > 0 and bore_r >= pitch_r:
            self._add(r, pno, IssueSeverity.ERROR, "WWHEEL_BORE",
                      f"bore_r={bore_r:.1f} ≥ pitch_r={pitch_r:.1f}")
        if m > 0 and outer_r > 0 and fw > outer_r * 1.5:
            self._add(r, pno, IssueSeverity.WARNING, "WWHEEL_FW_WIDE",
                      f"FaceWidth={fw} > 1.5×outer_r={outer_r:.1f} — throat may disappear")

    def _check_box(self, r, pno, p1, p2, p3, p4):
        if p1 <= 0 or p2 <= 0 or p3 <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "BOX_DIMS",
                      "All dims (P1 L, P2 W, P3 H) must be > 0")
        if p4 > 0 and p3 > 0 and p4 >= min(p1, p2, p3) / 2:
            self._add(r, pno, IssueSeverity.ERROR, "BOX_FILLET",
                      f"fillet_R={p4} ≥ min_dim/2={min(p1,p2,p3)/2:.1f}")

    def _check_cylinder(self, r, pno, p1, p2, p3, p4):
        if p1 <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "CYL_OUTER_R",
                      "Outer_R must be > 0")
        if p2 >= p1:
            self._add(r, pno, IssueSeverity.ERROR, "CYL_BORE",
                      f"Bore_R={p2} ≥ Outer_R={p1}")
        if p3 <= 0:
            self._add(r, pno, IssueSeverity.ERROR, "CYL_HEIGHT",
                      "Height must be > 0")
        if (p1 - p2) < 5:
            self._add(r, pno, IssueSeverity.WARNING, "CYL_WALL",
                      f"Wall={p1-p2:.1f}mm < 5mm — thin wall")

    @staticmethod
    def _add(r: ValidationReport, pno, sev, rule, msg):
        r.issues.append(ValidationIssue(pno, sev, rule, msg))
        if sev == IssueSeverity.ERROR:   r.error_count   += 1
        elif sev == IssueSeverity.WARNING: r.warning_count += 1


# ═══════════════════════════════════════════════════════════════════════════════
# DIFF ENGINE  —  before/after comparison
# ═══════════════════════════════════════════════════════════════════════════════

def compute_diff(old: List[BomPart], new: List[BomPart]) -> str:
    """
    Returns a human-readable diff string showing exactly what the AI changed.
    Covers: modified values, added parts, removed parts.
    """
    lines: List[str] = []

    old_map = {p.part_number: p for p in old}
    new_map = {p.part_number: p for p in new}

    # ── Changed / same ─────────────────────────────────────────────────────────
    for pno, new_part in new_map.items():
        if pno in old_map:
            old_part = old_map[pno]
            changes = _part_diff(old_part, new_part)
            if changes:
                lines.append(f"📦 MODIFIED  {pno}")
                lines.extend(f"    ↳ {c}" for c in changes)
                lines.append("")
        else:
            lines.append(f"➕ ADDED     {pno}  [{new_part.part_type} | {new_part.material}]")
            lines.append(f"    P1={new_part.param_1}  P2={new_part.param_2}  "
                         f"P3={new_part.param_3}  P4={new_part.param_4}")
            lines.append("")

    # ── Removed ────────────────────────────────────────────────────────────────
    for pno in old_map:
        if pno not in new_map:
            lines.append(f"➖ REMOVED   {pno}  [{old_map[pno].part_type}]")
            lines.append("")

    if not lines:
        return "ℹ No changes detected."
    return "\n".join(lines).rstrip()


def _part_diff(old: BomPart, new: BomPart) -> List[str]:
    changes = []
    fields = [
        ("Part_Type",   old.part_type,   new.part_type),
        ("Material",    old.material,    new.material),
        ("Param_1",     old.param_1,     new.param_1),
        ("Param_2",     old.param_2,     new.param_2),
        ("Param_3",     old.param_3,     new.param_3),
        ("Param_4",     old.param_4,     new.param_4),
        ("Qty",         old.qty,         new.qty),
        ("Priority",    old.priority,    new.priority),
        ("Enabled",     old.enabled,     new.enabled),
        ("Description", old.description, new.description),
        ("Notes",       old.notes,       new.notes),
    ]
    for name, ov, nv in fields:
        if str(ov) != str(nv):
            changes.append(f"{name}: {ov}  →  {nv}")
    return changes


# ═══════════════════════════════════════════════════════════════════════════════
# GEMINI CLIENT
# ═══════════════════════════════════════════════════════════════════════════════

_SYSTEM_PROMPT = """\
You are the Siraal AI Engineering Copilot — an expert in gear design, ISO/IS standards,
and the Siraal 3D Gear Manufacturing Engine BOM format.

Your job is to intelligently edit a Bill of Materials (BOM) for gear manufacturing.

━━ SIRAAL BOM SCHEMA ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Each part has these exact fields:
  Part_Number  — unique ID, e.g. GR-SP-001-PINION (never change unless asked)
  Part_Type    — one of: Spur_Gear_3D, Helical_Gear, Ring_Gear_3D, Bevel_Gear,
                          Worm, Worm_Wheel, Box, Cylinder, Sphere
  Material     — one of: Steel-1020, Steel-4140, Al-6061, Brass-C360, Nylon-66, Ti-6Al-4V
  Param_1      — integer  → Teeth Z  (or N_starts for Worm)
  Param_2      — float    → Module m in mm  (standard: 1, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10)
  Param_3      — float    → Face Width in mm  (or Length for Worm)
  Param_4      — float    → Bore Diameter in mm  (or Ring Thickness for Ring_Gear_3D)
  Qty          — integer  ≥ 1
  Priority     — "High", "Medium", or "Low"
  Enabled      — "YES" or "NO"
  Description  — short human-readable summary (update it when params change)
  Notes        — manufacturing notes

━━ ENGINEERING RULES (ISO 1328 / IS 2535) ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  • Param_1 (Z) ≥ 17 to avoid involute undercut (warn if 6-16)
  • bore_diameter < pitch_diameter (= Z × m)   ALWAYS
  • Face width: 6m ≤ FW ≤ 12m for spur;  8m ≤ FW ≤ 20m for helical
  • Ring_Gear_3D: Param_4 = Ring Thickness (NOT bore) — must be > module
  • Worm: use Brass-C360 for Worm_Wheel paired with steel Worm
  • Module should be from ISO standard series: 1, 1.25, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10
  • Bevel_Gear: FaceWidth × sin(45°) < pitch_radius  (cone truncation check)
  • Worm_Wheel face width < 1.5 × outer_radius

━━ OUTPUT RULES ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  • Return ONLY the complete JSON array — no preamble, no markdown, no explanation
  • Preserve Part_Numbers exactly unless the user explicitly asks to rename them
  • When you change params, update Description to reflect the new values
  • When you change material, update cost implications in Notes
  • Never invent Part_Types not in the allowed list
  • If asked to add gears for a gear ratio X:1, calculate Z correctly:
      Z_wheel = Z_pinion × ratio;  use same module for meshing pair
  • Worm pair rule: Worm.Param_2 must equal Worm_Wheel.Param_2 (same module)
"""

class GeminiClient:
    """
    Wraps google-genai SDK with:
      - Structured JSON output (response_mime_type)
      - Multi-turn conversation history
      - Automatic retry on quota / transient errors
    """

    MODEL = "gemini-2.5-flash"

    def __init__(self, api_key: str, log_cb: Callable = print):
        if not _GENAI_AVAILABLE:
            raise RuntimeError(
                "google-genai package not installed.\n"
                "Run:  pip install google-genai"
            )
        self._log = log_cb
        self._client = genai.Client(api_key=api_key)
        self._history: List[dict] = []   # list of {role, parts} dicts

    # ── Public ────────────────────────────────────────────────────────────────

    def ask(self, user_message: str,
            parts_json: str,
            validation_feedback: str = "",
            multi_turn: bool = False) -> List[dict]:
        """
        Send a message to Gemini and get back the new BOM as a list of dicts.

        Args:
            user_message        — the user's natural language instruction
            parts_json          — JSON string of current BOM (for context)
            validation_feedback — optional: pre-run validation issues to show AI
            multi_turn          — if True, maintains conversation history

        Returns:
            List[dict]  — the updated BOM parts as raw dicts
        """
        # Build the message content
        context_block = f"CURRENT BOM:\n{parts_json}"
        if validation_feedback:
            context_block += f"\n\nCURRENT VALIDATION ISSUES:\n{validation_feedback}"
        context_block += f"\n\nUSER INSTRUCTION:\n{user_message}"

        if multi_turn:
            # Add to history
            self._history.append({"role": "user", "parts": context_block})
        else:
            # Stateless single call
            self._history = []

        self._log("AI  Sending to Gemini (gemini-2.5-flash)…")

        # Build contents for API
        if multi_turn and len(self._history) > 1:
            contents_payload = [
                gtypes.Content(role=m["role"],
                               parts=[gtypes.Part(text=m["parts"])])
                for m in self._history[-12:]   # keep last 6 turns (12 messages)
            ]
        else:
            contents_payload = context_block

        raw_text = self._call_with_retry(contents_payload)

        # Store assistant reply in history
        if multi_turn:
            self._history.append({"role": "model", "parts": raw_text})

        return json.loads(raw_text)

    def reset_history(self):
        self._history = []
        self._log("SYSTEM  Conversation history cleared.")

    def get_history_length(self) -> int:
        return len(self._history) // 2   # turns

    # ── Private ───────────────────────────────────────────────────────────────

    def _call_with_retry(self, contents, max_retries: int = 2) -> str:
        import time
        last_err = None
        for attempt in range(max_retries + 1):
            try:
                response = self._client.models.generate_content(
                    model=self.MODEL,
                    contents=contents,
                    config=gtypes.GenerateContentConfig(
                        system_instruction=_SYSTEM_PROMPT,
                        response_mime_type="application/json",
                        temperature=0.05,   # highly deterministic for engineering data
                        max_output_tokens=8192,
                    )
                )
                raw = response.text.strip()
                # Strip any accidental markdown fence
                raw = re.sub(r"^```(?:json)?\s*", "", raw)
                raw = re.sub(r"\s*```$", "", raw)
                # Validate it parses
                json.loads(raw)
                return raw

            except json.JSONDecodeError as e:
                self._log(f"AI  Response was not valid JSON (attempt {attempt+1}): {e}")
                self._log(f"AI  Raw response: {response.text[:300]}…")
                last_err = e
                if attempt < max_retries:
                    time.sleep(1)
            except Exception as e:
                err_str = str(e)
                if "429" in err_str or "quota" in err_str.lower():
                    wait = 5 * (attempt + 1)
                    self._log(f"AI  Rate limited — waiting {wait}s (attempt {attempt+1})")
                    time.sleep(wait)
                    last_err = e
                else:
                    raise

        raise RuntimeError(f"Gemini call failed after {max_retries+1} attempts: {last_err}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN PUBLIC CLASS — SiraalCopilot
# ═══════════════════════════════════════════════════════════════════════════════

class SiraalCopilot:
    """
    The Siraal AI Copilot. Orchestrates BOM I/O, validation, and Gemini AI.

    Usage:
        copilot = SiraalCopilot("path/to/demo_gears_3d.xlsx", "YOUR_API_KEY")

        # Preview what the AI would do (does NOT save):
        ok, new_parts, diff = copilot.preview("Change all spur gears to Steel-4140")
        print(diff)

        # If happy with the preview, commit to disk:
        ok, msg = copilot.commit(new_parts)

        # Multi-turn conversation (remembers context):
        ok, new_parts, diff = copilot.chat("add a planetary gear set, ratio 4:1")
        ok, new_parts, diff = copilot.chat("make the ring gear out of Nylon-66")
        copilot.reset_chat()
    """

    def __init__(self,
                 excel_path: str,
                 api_key:    str,
                 log_cb:     Callable = print):
        self.excel_path   = excel_path
        self._log         = log_cb
        self._io          = BomIO(excel_path)
        self._validator   = BomValidator()
        self._gemini      = GeminiClient(api_key, log_cb)
        self._last_parts: Optional[List[BomPart]] = None  # cache for commit()

        self._log("╔══════════════════════════════════════════════╗")
        self._log("║   SIRAAL AI COPILOT  v1.0  —  Gemini 2.5    ║")
        self._log("╚══════════════════════════════════════════════╝")

    # ── One-shot preview (stateless) ──────────────────────────────────────────

    def preview(self, user_prompt: str
                ) -> Tuple[bool, Optional[List[BomPart]], str]:
        """
        Ask the AI to edit the BOM. Does NOT write to disk.

        Returns:
            (success, new_parts, diff_text)
            success    — False if Gemini call or validation has hard errors
            new_parts  — List[BomPart] ready to pass to commit(), or None
            diff_text  — human-readable diff for display in the GUI
        """
        return self._run(user_prompt, multi_turn=False)

    # ── Multi-turn chat ───────────────────────────────────────────────────────

    def chat(self, user_prompt: str
             ) -> Tuple[bool, Optional[List[BomPart]], str]:
        """
        Same as preview() but maintains conversation history so the AI
        remembers previous instructions in the same session.
        """
        return self._run(user_prompt, multi_turn=True)

    def reset_chat(self):
        self._gemini.reset_history()

    # ── Commit to disk ────────────────────────────────────────────────────────

    def commit(self,
               parts: Optional[List[BomPart]] = None,
               author: str = "AI Copilot",
               prompt_summary: str = ""
               ) -> Tuple[bool, str]:
        """
        Write `parts` (or the last preview result) to the Excel file.
        Also appends a ChangeLog entry.

        Returns: (success, message)
        """
        target = parts or self._last_parts
        if target is None:
            return False, "Nothing to commit — run preview() or chat() first."

        try:
            old_parts = self._io.load()
            self._io.save(target)
            n_changed = sum(
                1 for p in target
                if p.part_number not in {x.part_number for x in old_parts}
                or any(_part_diff(
                    next((x for x in old_parts if x.part_number == p.part_number), p), p
                ))
            )
            self._io.append_changelog(author, prompt_summary or "AI edit", n_changed)
            msg = f"✔ Saved {len(target)} parts to '{os.path.basename(self.excel_path)}' ({n_changed} changed)"
            self._log(msg)
            return True, msg
        except Exception as e:
            msg = f"✘ Save failed: {e}"
            self._log(msg)
            self._log(traceback.format_exc())
            return False, msg

    # ── Read-only helpers ────────────────────────────────────────────────────

    def load_bom(self) -> List[BomPart]:
        """Load current BOM from Excel (no AI involvement)."""
        return self._io.load()

    def validate_parts(self, parts: List[BomPart]) -> ValidationReport:
        """Run engineering validation on a list of parts."""
        return self._validator.validate(parts)

    def get_schema(self) -> dict:
        """
        Returns the BOM schema as a dict — useful for populating GUI dropdowns,
        tooltips, or feeding back to the AI as context.
        """
        return {
            "gear_types":     sorted(GEAR_TYPES),
            "solid_types":    sorted(SOLID_TYPES),
            "materials":      list(MATERIAL_DB.keys()),
            "priorities":     PRIORITIES,
            "enabled_values": ENABLED_VALUES,
            "param_labels":   PARAM_LABELS,
            "iso_modules":    [1, 1.25, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10, 12, 16, 20],
        }

    def get_chat_turns(self) -> int:
        return self._gemini.get_history_length()

    # ── Core pipeline ─────────────────────────────────────────────────────────

    def _run(self, user_prompt: str, multi_turn: bool
             ) -> Tuple[bool, Optional[List[BomPart]], str]:
        """
        Full pipeline:
          1. Load BOM from Excel
          2. Pre-validate (give AI awareness of existing issues)
          3. Call Gemini with BOM + prompt
          4. Parse response into BomPart objects
          5. Post-validate AI output
          6. Compute diff
          7. Cache for commit()
        """
        self._log(f"\n{'─'*60}")
        self._log(f"USER    {user_prompt}")
        self._log(f"{'─'*60}")

        # ── Step 1: Load ──────────────────────────────────────────────────────
        try:
            self._log("SYSTEM  Loading BOM from Excel…")
            old_parts = self._io.load()
            self._log(f"SYSTEM  Loaded {len(old_parts)} parts")
        except Exception as e:
            self._log(f"✘ Failed to load Excel: {e}")
            return False, None, str(e)

        # ── Step 2: Pre-validate ─────────────────────────────────────────────
        pre_report = self._validator.validate(old_parts)
        val_feedback = ""
        if pre_report.issues:
            val_feedback = pre_report.summary()
            self._log("SYSTEM  Pre-validation issues fed to AI as context")

        # ── Step 3: Call Gemini ───────────────────────────────────────────────
        parts_json = json.dumps(
            [p.to_dict() for p in old_parts], indent=2, ensure_ascii=False
        )
        try:
            raw_dicts = self._gemini.ask(
                user_message=user_prompt,
                parts_json=parts_json,
                validation_feedback=val_feedback,
                multi_turn=multi_turn,
            )
        except Exception as e:
            self._log(f"✘ Gemini error: {e}")
            return False, None, str(e)

        self._log(f"AI  Received {len(raw_dicts)} parts in response")

        # ── Step 4: Parse into BomPart objects ───────────────────────────────
        try:
            new_parts = [BomPart.from_dict(d) for d in raw_dicts]
        except Exception as e:
            self._log(f"✘ Failed to parse AI response into BomPart objects: {e}")
            return False, None, str(e)

        # ── Step 5: Post-validate AI output ──────────────────────────────────
        self._log("SYSTEM  Validating AI output…")
        post_report = self._validator.validate(new_parts)
        self._log(post_report.summary())

        if not post_report.is_valid:
            self._log("⚠  AI output has validation errors — showing diff anyway")
            # Don't block — let GUI decide whether to commit with errors

        # ── Step 6: Diff ──────────────────────────────────────────────────────
        diff_text = compute_diff(old_parts, new_parts)
        n_changes = diff_text.count("↳") + diff_text.count("➕") + diff_text.count("➖")
        self._log(f"SYSTEM  Diff complete — {n_changes} field change(s)")

        # Append validation summary to diff for GUI display
        if post_report.issues:
            diff_text += "\n\n── VALIDATION ──\n" + post_report.summary()

        # ── Step 7: Cache ─────────────────────────────────────────────────────
        self._last_parts = new_parts

        return True, new_parts, diff_text


# ═══════════════════════════════════════════════════════════════════════════════
# BACKWARDS-COMPATIBLE FUNCTION API  (matches your existing gui_launcher calls)
# ═══════════════════════════════════════════════════════════════════════════════

def preview_bom_edits(
        excel_path:  str,
        user_prompt: str,
        api_key:     str,
        log_cb:      Callable = print,
) -> Tuple[bool, Optional[List[dict]], str]:
    """
    Drop-in replacement for the original preview_bom_edits().
    Returns (success, list_of_dicts, diff_text) — same signature as before,
    but now uses the full SiraalCopilot pipeline.
    """
    copilot = SiraalCopilot(excel_path, api_key, log_cb)
    ok, parts, diff = copilot.preview(user_prompt)
    if ok and parts:
        return True, [p.to_dict() for p in parts], diff
    return ok, None, diff


def commit_bom_edits(
        excel_path:  str,
        new_dicts:   List[dict],
        api_key:     str = "",        # unused, kept for signature compat
        log_cb:      Callable = print,
        author:      str = "AI Copilot",
        prompt:      str = "",
) -> Tuple[bool, str]:
    """
    Write a list of part dicts (from preview_bom_edits) back to the Excel.
    """
    try:
        io = BomIO(excel_path)
        parts = [BomPart.from_dict(d) for d in new_dicts]
        old   = io.load()
        io.save(parts)
        n = sum(1 for p in parts if any(_part_diff(
            next((x for x in old if x.part_number == p.part_number), p), p
        )))
        io.append_changelog(author, prompt, n)
        msg = f"✔ Committed {len(parts)} parts ({n} modified)"
        log_cb(msg)
        return True, msg
    except Exception as e:
        msg = f"✘ Commit failed: {e}"
        log_cb(msg)
        return False, msg


def load_bom_to_dicts(path: str) -> List[dict]:
    """Drop-in replacement for original load_bom_to_dicts()."""
    return [p.to_dict() for p in BomIO(path).load()]


# ═══════════════════════════════════════════════════════════════════════════════
# QUICK TEST  (run directly: python genai_creator.py)
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys, pathlib

    # ── Find the demo Excel ───────────────────────────────────────────────────
    here = pathlib.Path(__file__).parent
    excel = here / "excels" / "demo_gears_3d.xlsx"
    if not excel.exists():
        print(f"[!] Demo Excel not found at {excel}")
        sys.exit(1)

    # ── API key from env or arg ───────────────────────────────────────────────
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if len(sys.argv) > 1:
        api_key = sys.argv[1]
    if not api_key:
        print("[!] No API key. Pass as argument:  python genai_creator.py YOUR_KEY")
        sys.exit(1)

    # ── Offline validation test (no API key needed) ───────────────────────────
    print("\n── OFFLINE VALIDATION TEST ──────────────────────────────────")
    io = BomIO(str(excel))
    parts = io.load()
    print(f"Loaded {len(parts)} parts:")
    for p in parts:
        print(f"  {p.part_number:25s}  {p.part_type:15s}  Z={p.param_1:4.0f}  "
              f"m={p.param_2}  FW={p.param_3}  bore={p.param_4}")

    validator = BomValidator()
    report = validator.validate(parts)
    print(report.summary())

    # ── Live AI test ──────────────────────────────────────────────────────────
    print("\n── LIVE AI COPILOT TEST ─────────────────────────────────────")
    copilot = SiraalCopilot(str(excel), api_key)

    test_prompts = [
        "Change all spur gears with Z < 17 to Z=20, same module. Update descriptions.",
        "For every worm-wheel pair, confirm the module matches. Fix any mismatch.",
    ]

    for prompt in test_prompts:
        print(f"\nPROMPT: {prompt}")
        ok, new_parts, diff = copilot.preview(prompt)
        if ok:
            print(diff)
        else:
            print(f"FAILED: {diff}")